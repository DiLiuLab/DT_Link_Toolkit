#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_diagramV2_5.py  --  Comprehensive diagram explorer for a single link.

New in V2.5 (harvesting is now the default generator)
-----------------------------------------------------
* Stage 1 no longer keeps ONE diagram per simplifier call.  ``backtrack_simplify``
  performs ``--backtrack-rounds`` complicate/re-simplify cycles internally and, until
  now, returned only the last strict improvement -- so a call with rounds=200 visited
  up to 200 minimal diagrams and discarded all but one.  The engine can now ARCHIVE
  every diagram that ties the minimum (link_engine_v4_0, ``collect_minimal=``), and
  ``generate_archive`` uses it.
* The effect is not marginal.  Calibrated against an exhaustive enumeration of every
  alternating knot of 10 crossings or fewer -- 196 knots, 509 minimal diagrams known
  exactly -- harvesting found 509 of 509, right on all 196.  The old chain sampler
  found 438 in one run and 463 pooled over three runs, and no amount of extra effort
  closed the gap: a 1600-round run found exactly what an 80-round run found.
* For the project's own 4BL link, 60 harvesting calls produced 6558 minimal codes that
  de-duplicate to the SAME four diagrams, so the four-diagram result now rests on a
  far larger sample.
* ``--legacy-chain`` restores the V2.3 one-diagram-per-round generator.  ``--rounds``
  now counts simplifier CALLS; far fewer are needed (6 sufficed for the calibration).

New in V2.3 (exploration strategy)
----------------------------------
* ``--reset-mode`` chooses WHERE a re-rooted chain restarts from.  Until now a reset
  always went back to the original diagram, which measurement showed is the weakest
  option: re-rooting at a fixed point re-explores the same neighbourhood.  Calibrated
  against an exhaustive enumeration (196 alternating knots, 509 minimal diagrams known
  exactly), a 1600-round chain re-rooted at the origin every 80 rounds found 438 --
  exactly what a plain 80-round chain found, and the IDENTICAL set on 180 of the 196.
  The three modes are now:
      origin   restart at the starting diagram            (the old behaviour)
      equal    restart at a uniformly random diagram among those found so far
      inverse  restart at one chosen with probability proportional to 1/(1+times used)
  ``equal`` gives every DISTINCT diagram the same chance of being a launch point, so
  rare diagrams get as much airtime as common ones; ``inverse`` goes further and
  prefers the least-used launch points, flattening the visit histogram the way
  Wang-Landau sampling does.  Both let the search creep outward to diagrams reachable
  only VIA another diagram, never directly from the start.
* The pool of restart candidates is keyed by ``canonical_key`` (the cheap WL-hash
  signature), so equivalent codes count once and the bookkeeping stays O(ms) per round.

New in V2.2 (all in the puncture / raw-grouping machinery)
----------------------------------------------------------
* The panels of the raw-grouping figure are grouped by EXACT marked-graph
  isomorphism instead of by comparing rendered crossing positions.  The old
  geometric key saw only the shadow -- the Tutte layout ignores over/under -- so
  two punctures whose drawings differ in which strand goes over produced the same
  key and were merged; the figure showed one panel FEWER than exists for three of
  the four project diagrams.  The new test needs no layout at all, which also
  removes one layout solve per candidate puncture.
* The panel grouping now also identifies a drawing with its mirror THROUGH the
  plane of the paper (every crossing flipped at once), matching the up-to-mirror
  convention `dedup` already applies to the classes themselves.  The two halves
  of the script previously disagreed about chirality.
* Punctured faces are pinned by IDENTITY rather than by their crossing-ID
  signature.  Two different faces can touch the same crossings -- the two
  triangles of the standard trefoil do -- and a signature cannot separate them.
* `--raw-verify` (on by default with `--raw-svg`) checks the figure two ways:
  that every face merged into a panel really does draw congruently to it, and,
  from the rendered geometry alone, that no two panels in a row are the same
  picture.
* New descriptive column **Plane drawings**: how many genuinely different plane
  pictures the diagram has, over ALL faces.  Intrinsic and puncture-independent,
  so unlike the 2-D numbers it is safe to compare across diagrams; not scored.
  Its face-orbit sizes cross-check the symmetry order by orbit-stabilizer.
* The figure's title no longer claims to show every plane drawing: it shows the
  ones draw_dt can produce, i.e. those from the largest-tie faces.
  enumerate_puncturing_dt.py enumerates all of them.

Pipeline
--------
1. GENERATE.  Starting from one signed DT code, run N rounds of simplification.
   Each round follows the *same* mechanism as strand_passage_guiV4_2.py:
   ``snappy.Link(dt) -> backtrack_simplify(mode='global') -> export new DT``.
   The randomized backtrack escapes local minima, so each round surfaces a
   (usually different) diagram of the *same* link; that new DT becomes the root
   of the next round.  N rounds => N+1 DT codes (1 initial + N simplified).

2. DEDUP.  Many of those DT codes are the *same diagram* written differently
   (cyclic re-labelling, direction reversal, component reordering, planar
   reflection/flip).  We collapse them to representatives using an exact
   signed-diagram isomorphism test (Weisfeiler-Lehman hash for bucketing +
   VF2 for confirmation) that preserves the over/under (chirality) pattern.

3. SCORE.  Each representative is scored with the score_diagram.py metric
   engine (combinatorial balance, planar-graph symmetry, 2-D Tutte energy,
   3-D sphere energy) and ranked by the composite quality.

4. REPORT.  Write an Excel workbook (one ranked row per representative with all
   metrics, plus a run_info sheet) and an SVG figure (2-D Tutte layout + 3-D
   sphere layout for each representative, captioned with its DT code and score).

Reproducibility: the generation step is resumable from a JSONL checkpoint, and the
checkpoint is the ONLY thing that reproduces a chain.  Round i is driven by a
per-round seed ``f(base_seed, i)``, but that does NOT make the chain repeatable:
spherogram's simplifier draws its moves from collections built with ``set()`` over
Crossing objects, which iterate in identity (memory-address) order, so seeding the
RNG does not pin the choices.  Measured: five calls to ``simplify_once`` with the
same root and the same seed return five different DT codes, in fresh processes with
PYTHONHASHSEED fixed.  Consequences:
  * re-running a chain explores a DIFFERENT sample of diagrams;
  * counts from one run are lower bounds, and two runs can differ in both
    directions;
  * a flat growth curve does not mean the search is complete -- measured on
    alternating knots against an exhaustive enumeration, 1600 rounds found no more
    diagrams than 80, while a second run of 80 rounds in a fresh process did.
Where an exact count matters, take the UNION over several runs (several processes),
not a longer run.

Requires SnapPy (``import snappy``); run under ``sage -python`` on the research
machine for the full toolchain.  Depends on score_diagram.py, link_engine_v4_0.py
and draw_dt_original_labelsV4_5.py sitting next to it.

Usage
-----
    python3 score_diagramV2_5.py                              # defaults: example DT, 24 calls
    python3 score_diagramV2_5.py --dt "DT: [...]" --rounds 24
    python3 score_diagramV2_5.py --xlsx out.xlsx --svg out.svg --json out.json
    # long runs can be chunked under a shell time limit:
    python3 score_diagramV2_5.py --generate-only --max-seconds 40   # repeat until done
"""

import argparse
import importlib.util
import json
import math
import os
import random
import re
import sys
import time

import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

import networkx as nx

_HERE = os.path.dirname(os.path.abspath(__file__))


def _find_base(filename):
    # Deliberately does NOT search old_scripts/.  That directory is excluded from
    # git, so a helper resolved there imports fine here and is simply absent in a
    # clone -- the failure is invisible locally and total for everyone else.
    for base in (_HERE, os.getcwd(), os.environ.get("DDOL_DIR", "")):
        if base and os.path.exists(os.path.join(base, filename)):
            return base
    return None


def _load_local(name, filename):
    """Import a sibling module by path, registered in sys.modules (dataclasses need it)."""
    base = _find_base(filename)
    if base is None:
        raise FileNotFoundError("Could not find %s next to score_diagramV2_5.py." % filename)
    if base not in sys.path:
        sys.path.insert(0, base)          # let intra-package `import ...` statements resolve
    spec = importlib.util.spec_from_file_location(name, os.path.join(base, filename))
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod


def _version_key(path):
    """Sort key for a versioned filename: its trailing V<major>_<minor>... as ints.

    Compares NUMERICALLY.  Plain text ordering would rank a future 'V10_0' below
    'V5_5', because '1' sorts before '5' character-wise.  Accepts every suffix
    spelling in this repo -- 'V5_5' (no separator), '_v4_0', '_V2_0'.  An
    unversioned file yields (), sorting below any versioned one.
    """
    stem = os.path.splitext(os.path.basename(path))[0]
    m = re.search(r"[_-]?[Vv](\d[A-Za-z0-9_]*)$", stem)
    return tuple(int(n) for n in re.findall(r"\d+", m.group(1))) if m else ()


def _find_draw_module():
    """Locate the current draw_dt_original_labels*.py, auto-adapting across version bumps
    (e.g. V4_5 -> V5_5 -> V10_0). Picks the highest-versioned file at the top level."""
    import glob
    for base in (_HERE, os.getcwd(), os.environ.get("DDOL_DIR", "")):
        if not base:
            continue
        matches = glob.glob(os.path.join(base, "draw_dt_original_labels*.py"))
        if matches:
            # basename breaks ties, so an equal-version pair resolves deterministically
            path = max(matches, key=lambda p: (_version_key(p), os.path.basename(p)))
            if base not in sys.path:
                sys.path.insert(0, base)
            return os.path.splitext(os.path.basename(path))[0], path
    raise FileNotFoundError(
        "Could not find draw_dt_original_labels*.py next to score_diagramV2_5.py.")


# Load the drawing helper under its real module name so link_engine's
# `import draw_dt_original_labelsV5_5 as D` reuses the same instance.
_draw_name, _draw_path = _find_draw_module()
_spec = importlib.util.spec_from_file_location(_draw_name, _draw_path)
DDOL = importlib.util.module_from_spec(_spec)
sys.modules[_draw_name] = DDOL
_spec.loader.exec_module(DDOL)
LE = _load_local("link_engine_v4_0", "link_engine_v4_0.py")
CDT2 = _load_local("canonical_dt_V2_0", "canonical_dt_V2_0.py")  # canonical form, combinatorial symmetry, and the rotation-system + eigenvalue point-group engine (sigma/i/Sn)

try:
    from scipy.optimize import linear_sum_assignment
    _HAVE_SCIPY = True
except Exception:  # pragma: no cover
    _HAVE_SCIPY = False


# =========================================================================== #
#  Metric engine (self-contained; the standalone score_diagram.py is archived
#  in old_scripts/ as the record-only original, so V2 does not import it).
# =========================================================================== #
def _cv(values):
    """Coefficient of variation (std / mean); 0 means perfectly uniform."""
    v = np.asarray(values, float)
    m = float(np.mean(v))
    if abs(m) < 1e-12:
        return 0.0
    return float(np.std(v) / m)


def _norm_entropy(counts):
    """Shannon entropy of a distribution, normalized to [0, 1]; 1 = perfectly even."""
    c = np.asarray(counts, float)
    s = float(np.sum(c))
    if s <= 0 or len(c) <= 1:
        return 1.0
    p = c / s
    p = p[p > 0]
    H = -float(np.sum(p * np.log(p)))
    return H / math.log(len(c))


def _assignment_residual(X, Y):
    """Mean matched distance between two equal-size point clouds (optimal bijection
    with SciPy, greedy otherwise)."""
    X = np.asarray(X, float)
    Y = np.asarray(Y, float)
    D = np.linalg.norm(X[:, None, :] - Y[None, :, :], axis=2)
    if _HAVE_SCIPY:
        r, c = linear_sum_assignment(D)
        return float(np.mean(D[r, c]))
    used, tot = set(), 0.0
    for i in range(D.shape[0]):
        for j in np.argsort(D[i]):
            if j not in used:
                used.add(j); tot += D[i, j]; break
    return tot / D.shape[0]


def _rotation_2d(theta):
    c, s = math.cos(theta), math.sin(theta)
    return np.array([[c, -s], [s, c]], float)


def _rotation_about_axis(axis, theta):
    a = np.asarray(axis, float)
    a = a / (np.linalg.norm(a) + 1e-15)
    x, y, z = a
    c, s = math.cos(theta), math.sin(theta)
    C = 1.0 - c
    return np.array([
        [c + x * x * C,     x * y * C - z * s, x * z * C + y * s],
        [y * x * C + z * s, c + y * y * C,     y * z * C - x * s],
        [z * x * C - y * s, z * y * C + x * s, c + z * z * C],
    ], float)


def combinatorial_metrics(model):
    crossings = model["crossings"]
    comp_positions = model["comp_positions"]
    comp_of = model["comp_of"]
    over_at = model["over_at"]
    n = len(crossings)
    C = len(comp_positions)

    visit_lengths = [len(cp) for cp in comp_positions]
    strand_cv = _cv(visit_lengths)
    strand_entropy = _norm_entropy(visit_lengths)
    strand_ratio = (max(visit_lengths) / min(visit_lengths)) if min(visit_lengths) else float("inf")

    L = np.zeros((C, C), int)
    n_self = 0
    for cr in crossings:
        ci = comp_of[cr["odd"]]
        cj = comp_of[cr["even"]]
        if ci == cj:
            n_self += 1
            L[ci, ci] += 1
        else:
            L[ci, cj] += 1
            L[cj, ci] += 1
    n_inter = n - n_self
    link_deg = [int(np.sum(L[i]) - L[i, i]) for i in range(C)]
    link_deg_cv = _cv(link_deg) if C > 1 else 0.0

    signs = [1 if cr["even_signed"] > 0 else -1 for cr in crossings]
    n_pos = sum(1 for s in signs if s > 0)
    n_neg = n - n_pos
    sign_imbalance = abs(n_pos - n_neg) / n

    flips = total = 0
    for cp in comp_positions:
        Lc = len(cp)
        if Lc < 2:
            continue
        for i in range(Lc):
            total += 1
            if over_at[cp[i]] != over_at[cp[(i + 1) % Lc]]:
                flips += 1
    alternating_frac = (flips / total) if total else 1.0

    return {
        "n_crossings": n, "n_components": C,
        "strand_visit_lengths": visit_lengths, "strand_length_cv": strand_cv,
        "strand_balance_entropy": strand_entropy, "strand_length_ratio": strand_ratio,
        "n_self_crossings": n_self, "n_inter_crossings": n_inter,
        "linking_matrix": L.tolist(), "linking_degrees": link_deg,
        "linking_degree_cv": link_deg_cv,
        "n_over_neg_convention_pos": n_pos, "n_neg": n_neg,
        "sign_imbalance": sign_imbalance, "alternating_fraction": alternating_frac,
    }


def _is_seg(node):
    return isinstance(node, tuple) and len(node) == 2 and node[0] == "seg"


def _crossing_of_corner(node):
    if isinstance(node, tuple) and not _is_seg(node) and isinstance(node[0], int):
        return node[0]
    return None


def planar_graph_metrics(model, G):
    ok, emb = nx.check_planarity(G)
    if not ok:
        raise RuntimeError("Diagram graph is not planar (DT may be non-realizable).")
    n = len(model["crossings"])
    faces = DDOL.planar_faces(emb)
    face_degrees = []
    for f in faces:
        segs = [x for x in f if _is_seg(x)]
        if segs:
            face_degrees.append(len(segs))
    face_degrees.sort()
    n_bigons = sum(1 for d in face_degrees if d == 2)
    return {
        "euler_faces": len(face_degrees), "euler_faces_expected": n + 2,
        "face_degrees": face_degrees,
        "face_degree_cv": _cv(face_degrees) if face_degrees else 0.0,
        "n_bigons": n_bigons,
        "automorphism_order": None,   # injected by score_diagram (robust canonical count)
        "n_plane_drawings": None,     # injected by score_diagram (needs the marked graph)
        "face_orbit_sizes": None,
        "orbit_check": None,
        "_emb": emb,
    }


def plane_drawing_count(model, G, emb):
    """How many genuinely different plane drawings the diagram has, in total.

    Every face is punctured, not only the largest, and the punctures are grouped
    by exact marked-graph isomorphism up to mirror -- so this counts the distinct
    pictures the diagram can present, and nothing about a particular layout.  It
    is INTRINSIC: unlike the 2-D Tutte numbers it does not depend on which face
    happens to be turned to the outside, which is what makes it safe to report
    alongside the scored metrics.

    Returns (count, orbit_sizes).  Fewer drawings means a more symmetric diagram:
    on the four project diagrams the count runs 3 / 6 / 8 / 9 for the rosette,
    Balanced, Offset and Lopsided clasps -- the same order as the composite.
    """
    recs = _diagram_face_records(model, emb)
    if not recs:
        return 0, []
    orbits = _puncture_orbits(model, G, recs, chiral_merge=True)
    return len(orbits), sorted(len(o) for o in orbits)


def check_face_orbits(orbit_sizes, automorphism_order, chiral_merge=True):
    """Orbit-stabilizer cross-check on the symmetry order.

    A group acting on the faces gives orbits whose sizes all DIVIDE its order, so
    this cross-checks two numbers computed by completely different routes:
    canonical_symmetry (canonical DT relabellings) for the order, and marked-graph
    isomorphism for the orbits.  A violation means one of them is wrong.

    Which group, though.  ``automorphism_order`` counts the symmetries that
    PRESERVE over/under.  When the orbits were built with the chirality flip
    merged in, the acting group is that one extended by the mirror through the
    plane of the paper -- an index-2 extension at most -- so the bound is 2|G|,
    not |G|.  Getting this wrong reports a failure on a perfectly good diagram:
    the orthogonal rosette has a face orbit of size 8 against |G| = 4.

    Returns "ok", "n/a", or a description of the failure.
    """
    if not orbit_sizes or not automorphism_order:
        return "n/a"
    bound = automorphism_order * (2 if chiral_merge else 1)
    bad = [k for k in orbit_sizes if bound % k]
    if bad:
        return ("FAIL: orbit size(s) %s do not divide %d (|G| = %d%s)"
                % (sorted(set(bad)), bound, automorphism_order,
                   ", doubled for the crossing flip" if chiral_merge else ""))
    return "ok"


def _signed_automorphism_order(model, G, cap=20000, time_limit=5.0):
    """Order of the diagram's symmetry group that PRESERVES over/under (the meaningful
    'same sequence' symmetry).  Nodes are labelled O/U/S and further coloured by
    Weisfeiler-Lehman refinement so VF2 only tries colour-preserving maps (automorphisms
    always preserve those colours, so the count is exact); a wall-clock guard in the main
    thread prevents pathological VF2 backtracking from hanging."""
    import threading
    H = nx.Graph(G)
    over_at = model["over_at"]
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        oo, oe = bool(over_at[cr["odd"]]), bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = "O" if oo else "U"
        for role in ("in_e", "out_e"):
            labels[(k, role)] = "O" if oe else "U"
    for node in H.nodes():
        labels.setdefault(node, "S")
    try:
        wl = nx.weisfeiler_lehman_subgraph_hashes(H, iterations=4)
        colors = {node: "%s|%s" % (labels[node], (h[-1] if h else "")) for node, h in wl.items()}
    except Exception:
        colors = labels
    nx.set_node_attributes(H, colors, "_c")
    nm = nx.algorithms.isomorphism.categorical_node_match("_c", None)
    GM = nx.algorithms.isomorphism.GraphMatcher(H, H, node_match=nm)
    count = {"n": 0}

    def _run():
        for _ in GM.isomorphisms_iter():
            count["n"] += 1
            if count["n"] >= cap:
                break

    timed_out = False
    if threading.current_thread() is threading.main_thread():
        import signal

        def _handler(signum, frame):
            raise TimeoutError()

        old = signal.signal(signal.SIGALRM, _handler)
        signal.setitimer(signal.ITIMER_REAL, time_limit)
        try:
            _run()
        except TimeoutError:
            timed_out = True
        finally:
            signal.setitimer(signal.ITIMER_REAL, 0)
            signal.signal(signal.SIGALRM, old)
    else:
        _run()
    return count["n"], count["n"] >= cap, timed_out


def _crossing_centers_2d(model, P):
    n = len(model["crossings"])
    centers = np.zeros((n, 2), float)
    counts = np.zeros(n, float)
    for node, xy in P.items():
        k = _crossing_of_corner(node)
        if k is not None:
            centers[k] += np.asarray(xy, float)
            counts[k] += 1
    counts[counts == 0] = 1
    return centers / counts[:, None]


def _diagram_arcs(model):
    pos_cross = model["pos_cross"]
    arcs = []
    for cp in model["comp_positions"]:
        Lc = len(cp)
        for i in range(Lc):
            arcs.append((pos_cross[cp[i]], pos_cross[cp[(i + 1) % Lc]]))
    return arcs


def _best_rotational_symmetry_2d(centers, kmax=8, thresh=0.12):
    X = np.asarray(centers, float)
    X = X - X.mean(axis=0)
    scale = float(np.sqrt(np.mean(np.sum(X ** 2, axis=1)))) + 1e-15
    # Report the score OF THE ORDER THAT WAS FOUND, and 0 when none was, exactly as
    # _best_rotational_symmetry_3d does.  The previous version carried a line that was
    # a no-op for every k != 2, so a diagram with no rotational symmetry at all still
    # reported "how close to 2-fold" -- the rosette came back as order 1 with score
    # 0.74, while its 3-D twin would have said 0.0 in the same situation.
    best_k, best_score = 1, 0.0
    for k in range(2, kmax + 1):
        Y = X @ _rotation_2d(2 * math.pi / k).T
        resid = _assignment_residual(X, Y) / scale
        if resid < thresh and k > best_k:
            best_k, best_score = k, max(0.0, 1.0 - resid)
    return best_k, best_score


def geometric_2d_metrics(model, G):
    P = DDOL.compute_positions(G, "tutte")
    centers = _crossing_centers_2d(model, P)
    span = float(np.sqrt(np.mean(np.sum((centers - centers.mean(0)) ** 2, axis=1)))) + 1e-15
    C = centers / span
    arcs = _diagram_arcs(model)
    lengths = np.array([np.linalg.norm(C[a] - C[b]) for a, b in arcs], float)
    edge_cv = _cv(lengths)
    dirichlet_norm = float(np.mean(lengths ** 2) / (np.mean(lengths) ** 2 + 1e-15))
    from collections import defaultdict
    nbrs = defaultdict(list)
    for a, b in arcs:
        nbrs[a].append(b)
        nbrs[b].append(a)
    devs = []
    for k in range(len(C)):
        others = nbrs.get(k, [])
        if len(others) < 2:
            continue
        angs = sorted(math.atan2(*(C[o] - C[k])[::-1]) for o in others)
        gaps = [(angs[(i + 1) % len(angs)] - angs[i]) % (2 * math.pi) for i in range(len(angs))]
        ideal = 2 * math.pi / len(gaps)
        devs.append(math.degrees(math.sqrt(np.mean([(g - ideal) ** 2 for g in gaps]))))
    angle_dev = float(np.mean(devs)) if devs else 0.0
    sym_k, sym_score = _best_rotational_symmetry_2d(centers)
    return {
        "edge_length_cv": edge_cv, "dirichlet_energy_norm": dirichlet_norm,
        "crossing_angle_rms_dev_deg": angle_dev,
        "sym2d_order": sym_k, "sym2d_score": sym_score,
        "_centers2d": centers, "_arcs": arcs,
    }


def _crossing_centers_3d(model, G):
    dirs = DDOL._kamada_3d_unit_directions(G)
    n = len(model["crossings"])
    centers = np.zeros((n, 3), float)
    counts = np.zeros(n, float)
    for node, d in dirs.items():
        k = _crossing_of_corner(node)
        if k is not None:
            centers[k] += np.asarray(d, float)
            counts[k] += 1
    counts[counts == 0] = 1
    centers = centers / counts[:, None]
    norms = np.linalg.norm(centers, axis=1, keepdims=True)
    norms[norms < 1e-12] = 1.0
    return centers / norms


def _riesz_energy(points, s=1.0):
    P = np.asarray(points, float)
    m = len(P)
    e = 0.0
    for i in range(m):
        for j in range(i + 1, m):
            d = np.linalg.norm(P[i] - P[j])
            if d > 1e-12:
                e += 1.0 / d ** s
    return float(e)


def _best_rotational_symmetry_3d(centers, kmax=6, thresh=0.12):
    X = np.asarray(centers, float)
    X = X - X.mean(axis=0)
    scale = float(np.sqrt(np.mean(np.sum(X ** 2, axis=1)))) + 1e-15
    _, _, Vt = np.linalg.svd(X - X.mean(0), full_matrices=False)
    axes = [Vt[i] for i in range(Vt.shape[0])]
    cen = X.mean(0)
    for i in range(len(X)):
        v = X[i] - cen
        if np.linalg.norm(v) > 1e-9:
            axes.append(v)
    best_k, best_score = 1, 0.0
    for axis in axes:
        for k in range(2, kmax + 1):
            Y = X @ _rotation_about_axis(axis, 2 * math.pi / k).T
            resid = _assignment_residual(X, Y) / scale
            score = max(0.0, 1.0 - resid)
            if resid < thresh and k > best_k:
                best_k, best_score = k, score
            elif k == best_k and score > best_score:
                best_score = score
    if best_k == 1:
        best_score = 0.0
    return best_k, best_score


def sphere_3d_metrics(model, G):
    centers = _crossing_centers_3d(model, G)
    n = len(centers)
    thomson = _riesz_energy(centers, s=1.0)
    ref = _riesz_energy(DDOL._fibonacci_sphere_directions(n), s=1.0)
    # The reference is a Fibonacci spiral -- evenly spread, but NOT the true
    # minimum-energy (Thomson) arrangement -- so a small, very even diagram can beat
    # it and the raw ratio can exceed 1.  Measured: Hopf 1.053, Borromean 1.011,
    # trefoil 1.002.  Every other quality is bounded in [0, 1] and the composite is
    # documented as a mean of qualities, so the ratio is CAPPED at 1 here: "at least
    # as evenly spread as the reference" is the best this term can say.  The raw
    # ratio is kept as sphere_spread_raw (JSON only) so nothing is lost.
    spread_raw = float(ref / thomson) if thomson > 0 else 1.0
    spread_quality = min(1.0, spread_raw)
    pos_cross = model["pos_cross"]
    comp_lengths, turning = [], []
    for cp in model["comp_positions"]:
        ks = [pos_cross[p] for p in cp]
        pts = centers[ks]
        m = len(pts)
        if m < 2:
            comp_lengths.append(0.0)
            continue
        clen = 0.0
        for i in range(m):
            clen += math.acos(float(np.clip(np.dot(pts[i], pts[(i + 1) % m]), -1, 1)))
        comp_lengths.append(clen)
        for i in range(m):
            t1 = pts[i] - pts[(i - 1) % m]
            t2 = pts[(i + 1) % m] - pts[i]
            n1, n2 = np.linalg.norm(t1), np.linalg.norm(t2)
            if n1 > 1e-9 and n2 > 1e-9:
                ang = math.acos(float(np.clip(np.dot(t1, t2) / (n1 * n2), -1, 1)))
                turning.append(ang ** 2)
    strand3d_cv = _cv(comp_lengths)
    bending = float(np.sum(turning))
    sym_k, sym_score = _best_rotational_symmetry_3d(centers)
    return {
        "thomson_energy": thomson, "thomson_reference": ref,
        "sphere_spread_quality": spread_quality, "sphere_spread_raw": spread_raw,
        "strand3d_length_cv": strand3d_cv,
        "bending_energy": bending, "sym3d_order": sym_k, "sym3d_score": sym_score,
        "_centers3d": centers,
    }


# Edit WEIGHTS to explore which properties should dominate the composite score.
WEIGHTS = {
    "strand_balance": 1.0,     # equal strand lengths across components
    "diagram_symmetry": 1.0,   # combinatorial (sign-aware) + 3-D geometric symmetry
    "face_regularity": 1.0,    # regular planar faces + few bigons (embedding property)
    "sphere_energy": 1.0,      # evenly spread crossings + low bending in 3-D
}


def _quality_scores(m):
    # The 2-D Tutte/geom2d numbers (edge-length CV, Dirichlet energy, crossing-angle
    # deviation, 2-D positional symmetry) depend on which face draw_dt turns to the
    # OUTSIDE (the 'puncture').  When several faces tie for the largest boundary that
    # choice is not intrinsic to the diagram, so those metrics are NOT scored -- they
    # are kept only as descriptive columns.  Every quality below is computed from
    # puncture-independent data: strand lengths, the sign-aware automorphism group,
    # the planar-embedding face spectrum, and the 3-D sphere layout.
    c, g, q3 = m["combinatorial"], m["graph"], m["sphere3d"]
    strand_balance = 0.5 * c["strand_balance_entropy"] + 0.5 * (1.0 / (1.0 + c["strand_length_cv"]))
    # Both terms are topological and sign-aware: the combinatorial automorphism
    # order and the loop-gated 3-D point-group order.  The dot-only sym3d_score is
    # deliberately NOT used -- it false-positives on low-symmetry diagrams and is
    # kept as a descriptive column only.
    aut = g["automorphism_order"]
    order3d = q3.get("point_group_order_3d", aut)
    sym_comb = 1.0 - 1.0 / max(1, aut)
    sym_3d = 1.0 - 1.0 / max(1, order3d)
    diagram_symmetry = np.mean([sym_comb, sym_3d])
    # 'face regularity' replaces the old puncture-dependent 'geometric strain':
    # face-degree CV and the bigon count are properties of the planar embedding
    # (the combinatorial map), independent of the outer-face choice.
    face_regularity = np.mean([
        1.0 / (1.0 + g["face_degree_cv"]),
        1.0 / (1.0 + g["n_bigons"]),
    ])
    sphere_energy = np.mean([q3["sphere_spread_quality"], 1.0 / (1.0 + q3["strand3d_length_cv"])])
    return {
        "strand_balance": float(strand_balance),
        "diagram_symmetry": float(diagram_symmetry),
        "face_regularity": float(face_regularity),
        "sphere_energy": float(sphere_energy),
    }


def score_diagram(dt_string, negative_even="over"):
    comps = DDOL.parse_dt(dt_string)
    model = DDOL.build_model(comps, negative_even=negative_even)
    G = DDOL.build_gadget_graph(model)
    m = {
        "dt": dt_string,
        "combinatorial": combinatorial_metrics(model),
        "graph": planar_graph_metrics(model, G),
        "geom2d": geometric_2d_metrics(model, G),
        "sphere3d": sphere_3d_metrics(model, G),
    }
    # robust, reproducible symmetry order + group (respecting over/under), from the
    # canonical DT relabellings; replaces the fragile VF2 automorphism enumeration.
    m["graph"]["automorphism_order"] = canonical_symmetry(dt_string)
    m["graph"]["symmetry_group"] = canonical_group(dt_string)
    # Distinct plane drawings: intrinsic (puncture-independent), so it is reported
    # alongside the scored metrics rather than greyed out with the 2-D numbers.
    # The orbit sizes then cross-check the symmetry order by orbit-stabilizer --
    # two independent computations that must agree.
    try:
        n_draw, orbit_sizes = plane_drawing_count(model, G, m["graph"]["_emb"])
        m["graph"]["n_plane_drawings"] = n_draw
        m["graph"]["face_orbit_sizes"] = orbit_sizes
        m["graph"]["orbit_check"] = check_face_orbits(
            orbit_sizes, m["graph"]["automorphism_order"])
    except Exception as exc:  # noqa: BLE001  -- never let a report kill a score
        m["graph"]["orbit_check"] = "error: %s" % exc
    # Loop-aware 3-D point-group order, computed here (before scoring) in this
    # diagram's own 3-D frame.  Feeds diagram_symmetry and also warms the shared
    # cache the descriptive "3D point group" column reads from.
    m["sphere3d"]["point_group_order_3d"] = _point_group_order_3d(
        dt_string, m["sphere3d"]["_centers3d"], m["graph"]["automorphism_order"])
    m["quality"] = _quality_scores(m)
    m["composite"] = float(sum(WEIGHTS[k] * m["quality"][k] for k in WEIGHTS) / sum(WEIGHTS.values()))
    return m


def _strip_private(m):
    out = {}
    for k, v in m.items():
        if isinstance(v, dict):
            out[k] = {kk: vv for kk, vv in v.items() if not kk.startswith("_")}
        else:
            out[k] = v
    return out


# --------------------------------------------------------------------------- #
#  1. Generation  (SnapPy global + backtrack, re-rooted each round)
# --------------------------------------------------------------------------- #
def simplify_once(dt, backtrack_rounds, backtrack_steps, seed):
    """One round: SnapPy global simplify with backtrack, return the new DT string.

    ``seed`` seeds ``random`` and ``numpy.random`` but does NOT determine the result:
    spherogram picks its moves out of identity-ordered sets, so repeated calls with the
    same root and seed return different codes.  See the module docstring."""
    import snappy
    random.seed(seed)
    try:
        np.random.seed(seed & 0x7FFFFFFF)
    except Exception:
        pass
    L = snappy.Link(dt)
    L = LE.backtrack_simplify(snappy, L, mode="global",
                              rounds=backtrack_rounds, steps=backtrack_steps)
    return LE.dt_to_string(LE.parse_dt_any(L.DT_code()))


_SUBPROC_SRC = """
import sys, random, importlib.util
import snappy
_spec = importlib.util.spec_from_file_location("LE", sys.argv[1])
LE = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(LE)
dt, br, bs, seed = sys.argv[2], int(sys.argv[3]), int(sys.argv[4]), int(sys.argv[5])
random.seed(seed)
try:
    import numpy as _np; _np.random.seed(seed & 0x7FFFFFFF)
except Exception:
    pass
L = snappy.Link(dt)
L = LE.backtrack_simplify(snappy, L, mode="global", rounds=br, steps=bs)
sys.stdout.write(LE.dt_to_string(LE.parse_dt_any(L.DT_code())))
"""


def simplify_once_subprocess(dt, backtrack_rounds, backtrack_steps, seed, timeout=600):
    """One round, run in a FRESH interpreter.

    spherogram picks its moves from identity-ordered collections, so the set of diagrams
    reachable from a given start depends on the process's allocation history.  Running a
    reset in a new process therefore explores a genuinely different part of the space --
    the effect that makes a union over separate runs larger than any single run.  Costs a
    process spawn plus a snappy import (~1-3 s), so it only pays at large --reset-every.
    Falls back to the in-process routine if the subprocess fails for any reason.
    """
    import subprocess
    try:
        out = subprocess.run([sys.executable, "-c", _SUBPROC_SRC, LE.__file__, dt,
                              str(backtrack_rounds), str(backtrack_steps), str(seed)],
                             capture_output=True, text=True, timeout=timeout)
        code = (out.stdout or "").strip()
        if code.startswith("DT:"):
            return code
    except Exception:  # noqa: BLE001
        pass
    return simplify_once(dt, backtrack_rounds, backtrack_steps, seed)


def reencode_same_diagram(dt, rng, tries=24):
    """A DIFFERENT DT string for the SAME diagram (new base points, directions, order).

    Correct by construction: choosing a base point and a direction per component, and an
    order of the components, is exactly a re-encoding, and _walk_to_dt returns None for
    any choice that is not a valid DT code.  No canonicalisation is needed to check it,
    which matters because canonicalising a many-component link is expensive.

    The point is that spherogram builds its internal objects -- and therefore its
    candidate ORDER -- from the code it is handed, so a fresh encoding of the same
    diagram explores differently.  Measured on K10a3: twelve encodings of one diagram,
    60 rounds each, found 4 diagrams individually but 9 between them.
    """
    tours, n = _diagram_tours(dt)
    variants = [_component_variants(t, False) for t in tours]
    order = list(range(len(tours)))
    for _ in range(tries):
        rng.shuffle(order)
        walk, bounds, off = [], [], 0
        for ci in order:
            walk.extend(rng.choice(variants[ci]))
            bounds.append((off, off + len(tours[ci])))
            off += len(tours[ci])
        tup = _walk_to_dt(walk, n, bounds)
        if tup is not None:
            cand = CDT2.fmt_dt(tup)
            if cand != dt:
                return cand
    return dt


def _read_checkpoint(path):
    chain = {}
    if path and os.path.exists(path):
        with open(path) as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                chain[int(rec["round"])] = rec["dt"]
    return [chain[i] for i in sorted(chain)] if chain else []


def _append_checkpoint(path, rnd, dt):
    if not path:
        return
    with open(path, "a") as fh:
        fh.write(json.dumps({"round": rnd, "dt": dt}) + "\n")


RESET_MODES = ("origin", "equal", "inverse")


def generate_chain(dt0, rounds, backtrack_rounds, backtrack_steps, base_seed,
                   checkpoint=None, max_seconds=None, reset_every=0, verbose=True,
                   reset_mode="origin", reset_reencode=False, reset_subprocess=False):
    """Return the list of DT strings [dt0, dt1, ..., dt_rounds]; resumable via checkpoint.

    NOT reproducible from ``base_seed`` -- see the module docstring.  Re-running gives
    a different sample; only the checkpoint replays a chain exactly.

    If ``reset_every > 0`` the chain is re-rooted at ``dt0`` after every
    ``reset_every`` rounds (i.e. rounds reset_every+1, 2*reset_every+1, ...
    start again from the original diagram).  This prevents the walk from getting
    trapped cycling among a few common minimal diagrams and re-seeds exploration
    from the canonical starting point.  Measured, on alternating knots with an
    exhaustive enumeration for ground truth: SHORT, FREQUENT restarts explore best.
    Against a plain 80-round chain, a second 80-round run at reset_every=10 added 25
    diagrams over 196 knots, while a 1600-round run at reset_every=80 added only 17 --
    twenty times the work for less gain.  Prefer several short runs to one long one."""
    chain = _read_checkpoint(checkpoint)
    if chain and verbose:
        print("  using EXISTING checkpoint '%s': %d round(s) already present (%d DT codes); "
              "generation resumes from there%s."
              % (checkpoint, len(chain) - 1, len(chain),
                 "" if (len(chain) - 1) < rounds else " (already at/beyond the requested rounds — "
                 "no new generation)"), flush=True)
    if not chain:
        chain = [dt0]
        _append_checkpoint(checkpoint, 0, dt0)
    elif chain[0] != dt0:
        raise ValueError("Checkpoint root DT does not match --dt; use a fresh --checkpoint.")

    if reset_mode not in RESET_MODES:
        raise ValueError("reset_mode must be one of %s" % (RESET_MODES,))

    # Pool of restart candidates, keyed by the cheap signature so that equivalent DT
    # strings count as ONE diagram.  value = [representative DT string, times used as root]
    rng = random.Random((int(base_seed) ^ 0x5EED) & 0x7FFFFFFF)
    pool, _keycache = {}, {}

    def _remember(dt):
        k = _keycache.get(dt)
        if k is None:
            try:
                k = canonical_key(dt)
            except Exception:  # noqa: BLE001  -- never let bookkeeping kill a run
                return
            _keycache[dt] = k
        pool.setdefault(k, [dt, 0])

    def _pick_root():
        """Where a reset restarts from, per reset_mode."""
        if reset_mode == "origin" or not pool:
            return dt0
        items = list(pool.values())
        if reset_mode == "equal":
            choice = rng.choice(items)                       # uniform over DISTINCT diagrams
        else:                                                # inverse: prefer least-used
            weights = [1.0 / (1.0 + it[1]) for it in items]
            choice = rng.choices(items, weights=weights, k=1)[0]
        choice[1] += 1
        return choice[0]

    for _dt in chain:                        # seed the pool from any resumed checkpoint
        _remember(_dt)

    t0 = time.time()
    while (len(chain) - 1) < rounds:
        i = len(chain)                       # next round index (1..rounds)
        seed = (int(base_seed) * 1000003 + i) & 0x7FFFFFFF
        root = chain[-1]
        is_reset = bool(reset_every) and i > 1 and ((i - 1) % int(reset_every) == 0)
        if is_reset:
            root = _pick_root()              # origin / equal / inverse
            if reset_reencode:               # same diagram, fresh encoding
                root = reencode_same_diagram(root, rng)
        if is_reset and reset_subprocess:    # fresh interpreter for this round
            dt_new = simplify_once_subprocess(root, backtrack_rounds, backtrack_steps, seed)
        else:
            dt_new = simplify_once(root, backtrack_rounds, backtrack_steps, seed)
        _remember(dt_new)
        chain.append(dt_new)
        _append_checkpoint(checkpoint, i, dt_new)
        if verbose and (i % 20 == 0 or i == rounds):
            print("  round %4d/%d  %.1fs" % (i, rounds, time.time() - t0), flush=True)
        if max_seconds and (time.time() - t0) >= max_seconds:
            break
    return chain


def generate_archive(dt0, calls, backtrack_rounds, backtrack_steps, base_seed,
                    checkpoint=None, max_seconds=None, verbose=True,
                    target_crossings=None):
    """Stage 1 by HARVESTING: keep every diagram that ties the minimum.

    Each call runs ``backtrack_rounds`` complicate/re-simplify cycles and archives every
    minimum-crossing diagram met along the way, rather than the single best one.  That is
    where the diagrams actually are: see the module docstring for the calibration.

    Returns the list of harvested DT strings (dt0 first).

    Each call restarts from a diagram at the BEST crossing count found so far (the
    starting code itself while it is still one of them).  The equal / inverse weightings
    that generate_chain offers are deliberately NOT used here: harvesting already sweeps
    a call's whole neighbourhood, and measured on a 24-crossing 5-component link the
    plain best-so-far rule gave 493 minimal codes in eight calls against 391 for uniform
    weighting -- the weighting only slows the descent by restarting from diagrams that
    are no longer minimal.
    """
    import snappy
    found = _read_checkpoint(checkpoint) or [dt0]
    bag = set(found)
    rng = random.Random((int(base_seed) ^ 0x5EED) & 0x7FFFFFFF)
    pool, _keycache = {}, {}
    _best = {"n": None}                      # running minimum crossing number
    _warned = {"origin": False}

    def _n_of(dt):
        try:
            return sum(len(t) for t in DDOL.parse_dt(dt))
        except Exception:  # noqa: BLE001
            return None

    n0 = target_crossings

    def _remember(dt):
        """Pool the diagram -- but only ever at the BEST crossing count seen.

        A restart above the running minimum wastes the whole call re-descending ground
        already covered.  Measured on a 24-crossing 5-component link before this guard:
        the roots handed to eight calls were 24, 22, 20, 24, 22, 22, 20, 22 -- half of
        them at or above the best already known, twice at the original code after a
        20-crossing diagram had been found.  So a better minimum PURGES the pool.
        """
        try:
            n = sum(len(t) for t in DDOL.parse_dt(dt))
        except Exception:  # noqa: BLE001
            return
        if _best["n"] is None or n < _best["n"]:
            _best["n"] = n
            pool.clear()
        elif n > _best["n"]:
            return
        k = _keycache.get(dt)
        if k is None:
            try:
                k = canonical_key(dt)
            except Exception:  # noqa: BLE001
                return
            _keycache[dt] = k
        pool.setdefault(k, [dt, 0])

    def _pick_root():
        """A diagram at the best crossing count so far; dt0 while it still qualifies."""
        if _best["n"] is None or _n_of(dt0) == _best["n"]:
            return dt0
        if not _warned["origin"]:
            _warned["origin"] = True
            if verbose:
                print("  [note] the starting code has %d crossings but %d has been reached; "
                      "restarting from the smaller diagrams instead"
                      % (_n_of(dt0), _best["n"]), flush=True)
        if not pool:
            return dt0
        return rng.choice(list(pool.values()))[0]

    if n0 is None:
        try:
            n0 = sum(len(t) for t in DDOL.parse_dt(dt0))
        except Exception:  # noqa: BLE001
            n0 = None
    _remember(dt0)

    t0 = time.time()
    for i in range(1, int(calls) + 1):
        root = dt0 if i == 1 else _pick_root()
        before = len(found)
        try:
            LE.backtrack_simplify(snappy, snappy.Link(root), mode="global",
                                  rounds=backtrack_rounds, steps=backtrack_steps,
                                  collect_minimal=bag)
        except Exception:  # noqa: BLE001
            pass
        seen_now = set(found)
        for dt in sorted(bag - seen_now):
            _remember(dt)                    # _remember itself keeps only the best
            found.append(dt)
            _append_checkpoint(checkpoint, len(found) - 1, dt)
        if verbose and (i % 5 == 0 or i == int(calls)):
            print("  call %3d/%d  %d codes harvested (+%d)  %.1fs"
                  % (i, calls, len(found), len(found) - before, time.time() - t0), flush=True)
        if max_seconds and (time.time() - t0) >= max_seconds:
            break

    # a harvesting call can dip BELOW the starting crossing number; keep the true minimum
    best = None
    for dt in found:
        try:
            n = sum(len(t) for t in DDOL.parse_dt(dt))
        except Exception:  # noqa: BLE001
            continue
        best = n if best is None else min(best, n)
    if best is not None:
        if verbose and best != _n_of(dt0):
            print("  minimum crossing number reached: %d (the given code has %d)"
                  % (best, _n_of(dt0)), flush=True)
        keep = []
        for dt in found:
            try:
                if sum(len(t) for t in DDOL.parse_dt(dt)) == best:
                    keep.append(dt)
            except Exception:  # noqa: BLE001
                pass
        if keep:
            found = keep
    return found


# --------------------------------------------------------------------------- #
#  2. Dedup  (exact signed-diagram isomorphism)
# --------------------------------------------------------------------------- #
def _iso_graph(dt):
    """Gadget graph with node labels encoding over/under; abstract-iso of this
    graph == same diagram up to rotation / reflection / relabelling / cyclic
    permutation / component reorder, while preserving the over/under pattern."""
    comps = DDOL.parse_dt(dt)
    model = DDOL.build_model(comps)
    G = DDOL.build_gadget_graph(model)
    over_at = model["over_at"]
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        over_o = bool(over_at[cr["odd"]])     # is the odd strand over here?
        over_e = bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = "O" if over_o else "U"
        for role in ("in_e", "out_e"):
            labels[(k, role)] = "O" if over_e else "U"
    for node in G.nodes():
        if node not in labels:
            labels[node] = "S"                # traversal-arc (segment) node
    nx.set_node_attributes(G, labels, "lab")
    return G, model


def canonical_key(dt):
    """Strong, cheap composite signature of a signed diagram:
    (Weisfeiler-Lehman hash of the over/under-labelled diagram graph,
     strand-length spectrum, planar face-degree spectrum).
    Isomorphic diagrams (rotation/reflection/relabel/cyclic-perm/component-reorder,
    over-under preserved) always share this signature, so it is a sound BUCKETER.

    It is NOT a decision procedure.  Collisions between genuinely different diagrams
    are common, not "astronomically unlikely" as this docstring used to claim: for a
    KNOT the strand spectrum is a single number, so two thirds of the signature carries
    no information, and K10a3's 12 minimal diagrams fall into just 4 buckets.  Callers
    must split each bucket with an exact test -- `dedup` does this via `_exact_iso`."""
    G, model = _iso_graph(dt)
    wl = nx.weisfeiler_lehman_graph_hash(G, node_attr="lab", iterations=5)
    strand = tuple(sorted(len(cp) for cp in model["comp_positions"]))
    ok, emb = nx.check_planarity(G)
    fdeg = []
    if ok:
        for f in DDOL.planar_faces(emb):
            segs = [x for x in f if _is_seg(x)]
            if segs:
                fdeg.append(len(segs))
    return (wl, strand, tuple(sorted(fdeg)))


def _exact_iso(dtA, dtB):
    GA, _ = _iso_graph(dtA)
    GB, _ = _iso_graph(dtB)
    if GA.number_of_nodes() != GB.number_of_nodes():
        return False
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    return nx.is_isomorphic(GA, GB, node_match=nm)


# --------------------------------------------------------------------------- #
#  Rigorous DT-native canonical form (authoritative "same diagram" test)
# --------------------------------------------------------------------------- #
# Two DT codes describe the SAME diagram iff, after re-deriving the DT under every
# choice of (component order, per-component base point, per-component traversal
# direction) [and optionally mirror = swapping over/under everywhere], their
# lexicographically-smallest valid DT codes are equal.  This is the exact meaning
# of "redundant due to symmetry / rotation / flipping / cyclic permutation", stated
# natively in DT terms and verifiable by hand.  It is O(product of component lengths)
# per diagram, so it is used for verification, not for bulk dedup.
def _diagram_tours(dt):
    m = DDOL.build_model(DDOL.parse_dt(dt))
    tours = [[(m["pos_cross"][p], bool(m["over_at"][p])) for p in cp]
             for cp in m["comp_positions"]]
    return tours, len(m["crossings"])


def _component_variants(tour, flip):
    base = [(c, (not o) if flip else o) for c, o in tour]
    L = len(base)
    out = []
    for seq in (base, base[::-1]):
        for s in range(L):
            out.append(seq[s:] + seq[:s])
    return out


def _walk_to_dt(walk, n, comp_bounds):
    slots = [[] for _ in range(n)]
    for i, (cr, ov) in enumerate(walk):
        slots[cr].append((i + 1, ov))
    signed = {}
    for lst in slots:
        if len(lst) != 2:
            return None
        (p1, o1), (p2, o2) = lst
        if (p1 & 1) == (p2 & 1):
            return None                        # not a valid DT for this base point
        if p1 & 1:
            oddp, evenp, ev = p1, p2, o2
        else:
            oddp, evenp, ev = p2, p1, o1
        signed[oddp] = (-evenp if ev else evenp)   # convention: negative even = even over
    tup = []
    for lo, hi in comp_bounds:
        odds = sorted(p for p in signed if (p & 1) and lo < p <= hi)
        tup.append(tuple(signed[p] for p in odds))
    return tuple(tup)


def canonical_dt(dt, allow_flip=True, return_symmetry=False):
    """Lexicographically minimal valid signed DT over all relabellings (and, if
    allow_flip, mirror).  Equal canonical_dt  <=>  same diagram.  When return_symmetry
    is True, also return how many valid relabellings reproduce that minimal code, which
    is the diagram's symmetry order (the count of DT re-encodings that coincide with the
    canonical form) - a robust, VF2-free symmetry measure."""
    from itertools import permutations, product
    tours, n = _diagram_tours(dt)
    C = len(tours)
    best = None
    n_min = 0
    for flip in ((False, True) if allow_flip else (False,)):
        var = [_component_variants(tours[ci], flip) for ci in range(C)]
        for perm in permutations(range(C)):
            bounds, off = [], 0
            for ci in perm:
                bounds.append((off, off + len(tours[ci])))
                off += len(tours[ci])
            for choice in product(*[var[ci] for ci in perm]):
                walk = []
                for seq in choice:
                    walk.extend(seq)
                tup = _walk_to_dt(walk, n, bounds)
                if tup is None:
                    continue
                if best is None or tup < best:
                    best, n_min = tup, 1
                elif tup == best:
                    n_min += 1
    return (best, n_min) if return_symmetry else best


_ISO_GRAPH_CACHE = {}


def _iso_graph_cached(dt):
    """`_iso_graph` memoised on the DT string -- dedup asks for the same graphs a lot."""
    g = _ISO_GRAPH_CACHE.get(dt)
    if g is None:
        g = _iso_graph(dt)[0]
        _ISO_GRAPH_CACHE[dt] = g
    return g


def _split_signature_bucket(strings):
    """Split one canonical_key bucket into genuinely distinct diagrams, UP TO MIRROR.

    The signature only buckets (see `canonical_key`); this is the exact step.  Each
    string is compared by labelled-graph VF2 against one representative per subgroup,
    directly and after mirroring, so a diagram and its mirror image stay together --
    which is what the up-to-mirror diagram count means.

    VF2 on these graphs is milliseconds, so this is cheap; it is NOT the factorial
    canonicalisation of `_mirror_canonical` and is not subject to the cost guard.
    """
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    subs = []                                  # [[rep_graph, rep_mirror_graph_or_None, [dt...]]]
    for dt in strings:
        G = _iso_graph_cached(dt)
        placed = False
        for sub in subs:
            if G.number_of_nodes() != sub[0].number_of_nodes():
                continue
            if nx.is_isomorphic(G, sub[0], node_match=nm):
                sub[2].append(dt); placed = True; break
            if sub[1] is None:
                try:
                    sub[1] = _iso_graph_cached(_mirror_dt(sub[2][0]))
                except Exception:  # noqa: BLE001
                    sub[1] = False
            if sub[1] and nx.is_isomorphic(G, sub[1], node_match=nm):
                sub[2].append(dt); placed = True; break
        if not placed:
            subs.append([G, None, [dt]])
    return [sub[2] for sub in subs]


def dedup(chain):
    """Collapse identical diagrams.  Returns a list of class dicts sorted by first
    appearance.

    Two stages: the cheap `canonical_key` signature buckets the distinct DT strings,
    then `_split_signature_bucket` splits each bucket by exact labelled-graph VF2.
    The second stage is essential -- the signature collides routinely (see
    `canonical_key`), and without it different diagrams are merged and the diagram
    count comes out too low."""
    # 1. group member round-indices by exact DT string, preserving first-seen order
    str_members, str_first = {}, {}
    for idx, dt in enumerate(chain):
        if dt not in str_members:
            str_members[dt] = []
            str_first[dt] = idx
        str_members[dt].append(idx)
    unique = sorted(str_members, key=lambda s: str_first[s])

    # 2. signature once per distinct string; group by signature.
    #    The signature (WL hash of the over/under-labelled diagram graph + strand-length
    #    spectrum + face-degree spectrum) is an isomorphism invariant, so identical
    #    diagrams always land together; a collision between genuinely different diagrams
    #    is astronomically unlikely and can be ruled out with --verify (exact VF2) or
    #    canonical_dt().
    groups = {}
    str_strings = {}
    for dt in unique:
        sig = canonical_key(dt)
        groups.setdefault(sig, []).append(dt)

    # 3. the signature only BUCKETS -- split each bucket by exact labelled-graph VF2.
    #    Without this step a signature collision silently merges different diagrams:
    #    K10a3's 12 minimal diagrams share only 4 signatures.
    classes = []
    for sig, strings in groups.items():
        for part in (_split_signature_bucket(strings) if len(strings) > 1 else [strings]):
            members = sorted(m for s in part for m in str_members[s])
            classes.append({
                "rep_dt": part[0],             # earliest first-seen string in the class
                "strings": part,
                "members": members,
                "multiplicity": len(members),
                "rep_round": min(members),
                "n_distinct_strings": len(part),
                "sig": sig,
            })
    classes.sort(key=lambda c: c["rep_round"])

    # UP-TO-MIRROR MERGE: collapse classes that are mirror images of each other.
    # The whole V2_0 framework (canonical form + symmetry) works up to mirror, so the
    # diagram COUNT should too.  The two Offset clasps are mirror images, so this takes
    # the set from 5 to 4 (Offset a/b become one entry, an enantiomeric pair).  Keyed by
    # the up-to-mirror canonical DT; the earliest-seen class leads each merged group and
    # carries that canonical as its representative (so scores/names are deterministic).
    by_mirror = {}
    for c in classes:
        mk = _mirror_canonical(c["rep_dt"])
        by_mirror.setdefault(mk, []).append(c)
    merged = []
    for mk, cs in by_mirror.items():
        cs.sort(key=lambda c: c["rep_round"])
        base = cs[0]
        for other in cs[1:]:
            base["strings"] = base["strings"] + other["strings"]
            base["members"] = sorted(set(base["members"]) | set(other["members"]))
        base["multiplicity"] = len(base["members"])
        base["rep_round"] = min(base["members"])
        base["n_distinct_strings"] = len(base["strings"])
        base["mirror_canonical"] = mk
        base["mirror_merged"] = len(cs) > 1
        # Keep EVERY merged signature, not just the leader's.  A mirror-merged class
        # contains members whose signature is the partner's (WL/face signatures are
        # over/under-preserving, so a diagram and its mirror have different ones), and
        # check_sampled looks members up by signature: dropping them made it answer
        # "NOT found" for diagrams the run had demonstrably sampled.
        base["sigs"] = [c["sig"] for c in cs]
        merged.append(base)
    merged.sort(key=lambda c: c["rep_round"])
    classes = merged
    for j, c in enumerate(classes, start=1):
        c["rep_id"] = j
    return classes


def _mirror_dt(dt):
    """The mirror-image DT code: negate every (even) entry, i.e. switch every crossing's
    over/under.  Used so verification matches the UP-TO-MIRROR de-duplication."""
    comps = CDT2.parse_dt(dt)
    return CDT2.fmt_dt(tuple(tuple(-x for x in c) for c in comps))


def verify_classes(classes, sample=25):
    """Confidence check: within each class, run exact labelled-graph VF2 between the
    representative and up to `sample` other distinct DT strings (0 = all).  Returns a
    per-class report; a False means the fast signature merged non-isomorphic diagrams.

    Because de-duplication now works UP TO MIRROR (a diagram and its mirror image share
    a class -- e.g. the two Offset clasps), a member counts as consistent if it is VF2-
    isomorphic to the representative EITHER directly OR after mirroring (negating the DT).
    Without this, VF2 -- which is over/under-preserving and does NOT see mirror images as
    isomorphic -- falsely reports a MERGE ERROR on any mirror-merged class."""
    import random as _r
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    report = []
    for c in classes:
        others = [s for s in c["strings"] if s != c["rep_dt"]]
        if sample and len(others) > sample:
            others = _r.Random(0).sample(others, sample)
        Grep, _ = _iso_graph(c["rep_dt"])

        def _iso_upto_mirror(o):
            if nx.is_isomorphic(Grep, _iso_graph(o)[0], node_match=nm):
                return True
            try:                                   # allow the mirror image (up-to-mirror dedup)
                return nx.is_isomorphic(Grep, _iso_graph(_mirror_dt(o))[0], node_match=nm)
            except Exception:  # noqa: BLE001
                return False

        ok = all(_iso_upto_mirror(o) for o in others)
        report.append({"rep_id": c["rep_id"], "checked": len(others), "all_isomorphic": ok})
    return report


def check_sampled(classes, queries):
    """For each query DT code, report whether an equivalent diagram was sampled
    (i.e. its signature matches one of the representative classes).

    Matching is UP TO MIRROR, like dedup itself: a class merged from an enantiomeric
    pair answers to both signatures (see the "sigs" list built in dedup), and a query
    is also tried mirrored.  Both are needed -- the WL/face signature preserves
    over/under, so a diagram and its mirror image never share one."""
    sigmap = {}
    for c in classes:
        for sig in (c.get("sigs") or [c["sig"]]):
            sigmap.setdefault(sig, c)
    out = []
    for q in queries:
        try:
            k = canonical_key(q)
        except Exception as exc:  # noqa: BLE001
            out.append({"dt": q, "sampled": False, "error": str(exc)})
            continue
        c = sigmap.get(k)
        if c is None:                       # try the mirror image (up-to-mirror dedup)
            try:
                c = sigmap.get(canonical_key(_mirror_dt(q)))
            except Exception:  # noqa: BLE001
                c = None
        out.append({
            "dt": q,
            "sampled": c is not None,
            "matches_rep_id": (c["rep_id"] if c else None),
            "multiplicity": (c["multiplicity"] if c else 0),
            "first_round": (c["rep_round"] if c else None),
        })
    return out


# --------------------------------------------------------------------------- #
#  3. Score + rank
# --------------------------------------------------------------------------- #
def _jones(dt):
    try:
        import snappy
        return str(snappy.Link(dt).jones_polynomial())
    except Exception:
        return "n/a (needs Sage)"


def _linking_matrix(L):
    """Pairwise linking numbers of a SnapPy/spherogram Link, WITHOUT Sage.

    ``Link.linking_matrix()`` is decorated ``@sage_method`` in spherogram, so it
    raises SageNotAvailable headless even though its body is plain Python over the
    crossing signs.  Recomputing it here is what makes _linking_fp live up to its
    name: previously every headless run silently produced no fingerprint at all and
    the same-link check reported "0 fingerprints (expected 1)" on a perfectly good
    result.  lk(i, j) = half the signed count of crossings between components i and j.
    """
    comps = L.link_components
    n = len(comps)
    M = [[0] * n for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            tot = 0.0
            for c in L.crossings:
                a = sum(1 for x in comps[i] if x[0] is c)
                b = sum(1 for x in comps[j] if x[0] is c)
                if a == 1 and b == 1:
                    tot += 0.5 * c.sign
            M[i][j] = M[j][i] = int(tot)
    return M


def _linking_fp(dt):
    """Headless same-link fingerprint: sorted off-diagonal linking numbers."""
    try:
        import snappy
        L = snappy.Link(dt)
        try:
            M = L.linking_matrix()          # fast path when running under Sage
        except Exception:                   # noqa: BLE001  -- SageNotAvailable, headless
            M = _linking_matrix(L)
        vals = sorted(int(M[i][j]) for i in range(len(M)) for j in range(len(M)) if i < j)
        return tuple(vals)
    except Exception:
        return None


_CANON_CACHE = None
_CANON_CACHE_PATH = os.environ.get("CANON_CACHE", "canonical_cache.json")


def _canonical_entry(dt, allow_flip=False):
    """Return {'dt': canonical string, 'sym': symmetry order}, memoized on disk so the
    (somewhat expensive) canonicalization is computed once per diagram per run."""
    global _CANON_CACHE
    if _CANON_CACHE is None:
        _CANON_CACHE = {}
        if os.path.exists(_CANON_CACHE_PATH):
            try:
                with open(_CANON_CACHE_PATH) as fh:
                    _CANON_CACHE = json.load(fh)
            except Exception:
                _CANON_CACHE = {}
    key = "N:" + dt
    entry = _CANON_CACHE.get(key)
    if isinstance(entry, dict) and all(k in entry for k in ("dt", "sym", "group")):
        return entry
    if _too_complex(dt, "symmetry order"):
        # Not cached to disk: it is a fallback, not a computed answer.
        return {"dt": dt, "sym": 1, "group": "n/a (too complex)"}
    # canonical form + symmetry order + symmetry GROUP, via the shared canonical_dt module
    res = CDT2.analyze(dt)
    entry = {"dt": res["canonical"], "sym": int(res["symmetry_order"]),
             "group": _short_group(res.get("element_orders") or [1] * res["symmetry_order"])}
    _CANON_CACHE[key] = entry
    _CANON_CACHE["N:" + entry["dt"]] = entry            # canonical maps to itself
    try:
        with open(_CANON_CACHE_PATH, "w") as fh:
            json.dump(_CANON_CACHE, fh)
    except Exception:
        pass
    return entry


def _short_group(orders):
    """Compact symmetry-group name from the multiset of element orders."""
    n = len(orders)
    if n <= 1:
        return "C1"
    if max(orders) == n:
        return "C%d" % n                                # cyclic (has an order-n element)
    if n == 4 and sorted(orders) == [1, 2, 2, 2]:
        return "C2xC2"                                  # Klein four-group
    if n == 6 and sorted(orders) == [1, 2, 2, 2, 3, 3]:
        return "D3"                                     # dihedral of order 6
    return "order-%d" % n


def canonical_dt_string(dt, allow_flip=False):
    if allow_flip:
        return _mirror_canonical(dt)
    return _canonical_entry(dt)["dt"]


# Canonicalisation enumerates C! x prod(2*Li) relabellings (Li = positions in component
# i), which is factorial in the component count.  4BL sits at 884 736 and takes ~5 s; a
# 5-component 22-crossing link needs 1.4e8 and a 28-crossing one 7.6e8, i.e. hours per
# class.  Past the limit the exact form is skipped and the cheap signature is used
# instead -- see _canonical_cost / _too_complex.
CANONICAL_COST_LIMIT = 5_000_000
_COMPLEXITY_WARNED = set()


def _canonical_cost(dt):
    """How many relabellings the exact canonical form would have to enumerate."""
    try:
        comps = DDOL.parse_dt(dt)
    except Exception:  # noqa: BLE001
        return None
    cost = math.factorial(len(comps))
    for c in comps:
        cost *= 4 * len(c)          # 2 * (positions in the component) = 2 * (2 * entries)
    return cost


def _too_complex(dt, what="canonical form"):
    """True when the exact canonicalisation is out of reach; warns once per run."""
    if not CANONICAL_COST_LIMIT:
        return False
    cost = _canonical_cost(dt)
    if cost is None or cost <= CANONICAL_COST_LIMIT:
        return False
    key = "".join(str(dt).split())
    if key not in _COMPLEXITY_WARNED:
        _COMPLEXITY_WARNED.add(key)
        try:
            comps = DDOL.parse_dt(dt)
            print("  [complexity] %d components / %d crossings would need %.3g relabellings "
                  "for the exact %s (limit %.3g).  Falling back to the signature; grouping is "
                  "still exact up to the WL hash, but the mirror merge and the symmetry order "
                  "are SKIPPED for this diagram."
                  % (len(comps), sum(len(c) for c in comps), cost, what, CANONICAL_COST_LIMIT),
                  flush=True)
        except Exception:  # noqa: BLE001
            pass
    return True


_MIRROR_CANON_CACHE = {}


def _mirror_canonical(dt):
    """UP-TO-MIRROR canonical DT string (V2_0, allow_flip=True): a diagram and its
    mirror image collapse to one code.  Used to merge mirror-image classes (the two
    Offset clasps are mirror images, so up to mirror there are 4 diagrams, not 5)."""
    key = "".join(str(dt).split())
    if key in _MIRROR_CANON_CACHE:
        return _MIRROR_CANON_CACHE[key]
    if _too_complex(dt, "canonical form"):
        # Return the code unchanged: still a VALID DT string (everything downstream scores
        # and draws it), just not canonical -- so mirror-image classes will not merge.
        _MIRROR_CANON_CACHE[key] = dt
        return dt
    try:
        comps = [tuple(c) for c in CDT2.parse_dt(dt)]
        canon, _, _ = CDT2.canonicalize(comps, allow_flip=True)
        s = CDT2.fmt_dt(canon)
    except Exception:  # noqa: BLE001
        s = _canonical_entry(dt)["dt"]
    _MIRROR_CANON_CACHE[key] = s
    return s


def canonical_symmetry(dt, allow_flip=False):
    return _canonical_entry(dt)["sym"]


def canonical_group(dt):
    return _canonical_entry(dt)["group"]


def score_representatives(classes):
    """Score each class on its CANONICAL DT code so the geometric/energy metrics are
    reproducible and independent of which sampled DT string first represented the class
    (the layout of two isomorphic-but-differently-labelled codes can differ, which would
    otherwise make the composite score wobble between runs)."""
    scored = []
    for c in classes:
        cdt = c.get("mirror_canonical") or _mirror_canonical(c["rep_dt"])  # up-to-mirror canonical
        c["canonical_dt"] = cdt
        m = score_diagram(cdt)            # m["dt"] == up-to-mirror canonical form
        m["_class"] = c
        m["jones"] = _jones(cdt)
        m["linking_fp"] = _linking_fp(cdt)
        scored.append(m)
    scored.sort(key=lambda m: m["composite"], reverse=True)
    for rank, m in enumerate(scored, start=1):
        m["rank"] = rank
    return scored


# --------------------------------------------------------------------------- #
#  Descriptive names for the five distinct diagrams of the target link
# --------------------------------------------------------------------------- #
# Structural finding (see the comparative report): four of the five diagrams
# contain a Bing-double "clasp" of two components -- a pair that is geometrically
# clasped yet has linking number 0.  In every diagram two "frame" components carry
# 3 crossings each, and the remaining two components split 8 crossings between
# them; the split fixes the clasp's balance (4-4, 5-3, or 6-2).  The fifth diagram
# spreads the same 8 crossings symmetrically ("orthogonally") with no localized
# clasp, which is why it alone reaches C2xC2 symmetry.  Names are keyed to the
# canonical DT code (whitespace-insensitive) so they are stable across runs and
# independent of the tie-order of the two 5-3 variants.
_DIAGRAM_NAMES = {
    "DT:[(-28,-26,-24,-22),(-20,-8,-4),(-6,12,-2),(-10,-16,-14,-18)]":
        ("Orthogonal rosette", "clasp-free; 8 crossings spread symmetrically (4-4); 3D point group C2v"),
    "DT:[(-28,-22,10,18),(-8,24,-20),(12,6,-26,-2),(4,16,-14)]":
        ("Balanced clasp", "Bing-double clasp, balanced 4-4 split; 3D point group Ci (inversion centre)"),
    # The two 5-3 Offset clasps are mirror images of each other, so up to mirror they
    # are ONE diagram (an enantiomeric pair).  Both codes map to the same name; the
    # merged class is represented by the up-to-mirror canonical (the second code).
    "DT:[(-28,-26,8),(14,22,-2),(-24,-20,-12),(-16,6,4,-18,-10)]":
        ("Offset clasp", "Bing-double clasp, offset 5-3 split (mirror-image pair a/b); 3D point group Cs (mirror plane)"),
    "DT:[(-28,-26,-22,12,-20),(-2,24,-18),(-4,-10,14),(8,6,-16)]":
        ("Offset clasp", "Bing-double clasp, offset 5-3 split (mirror-image pair a/b); 3D point group Cs (mirror plane)"),
    "DT:[(-28,-26,-22,14,16,20),(-4,24,-2),(-6,12),(10,8,-18)]":
        ("Lopsided clasp", "Bing-double clasp, lopsided 6-2 split; 3D point group C1"),
}


def _diagram_name(dt):
    """Return (name, clasp-structure) for a canonical DT, or ('','') if unknown."""
    return _DIAGRAM_NAMES.get("".join(str(dt).split()), ("", ""))


# --------------------------------------------------------------------------- #
#  4a. Excel report
# --------------------------------------------------------------------------- #
def _cols():
    # (header, path-getter) ; getter takes the scored metrics dict m
    def g(*keys):
        def _get(m):
            d = m
            for k in keys:
                d = d[k]
            return d
        return _get
    return [
        ("Rank", lambda m: m["rank"]),
        ("Rep ID", lambda m: m["_class"]["rep_id"]),
        ("Diagram name", lambda m: _diagram_name(m["dt"])[0]),
        ("Clasp structure", lambda m: _diagram_name(m["dt"])[1]),
        ("Canonical DT code", lambda m: m["dt"]),
        ("Multiplicity", lambda m: m["_class"]["multiplicity"]),
        ("First round", lambda m: m["_class"]["rep_round"]),
        ("Crossings", g("combinatorial", "n_crossings")),
        ("Components", g("combinatorial", "n_components")),
        ("Strand lengths", lambda m: str(m["combinatorial"]["strand_visit_lengths"])),
        ("Strand length CV", g("combinatorial", "strand_length_cv")),
        ("Strand balance entropy", g("combinatorial", "strand_balance_entropy")),
        ("Strand length max/min", g("combinatorial", "strand_length_ratio")),
        ("Self crossings", g("combinatorial", "n_self_crossings")),
        ("Inter crossings", g("combinatorial", "n_inter_crossings")),
        ("Linking degree CV", g("combinatorial", "linking_degree_cv")),
        ("Alternating fraction", g("combinatorial", "alternating_fraction")),
        ("Over/under imbalance", g("combinatorial", "sign_imbalance")),
        ("Faces (n+2)", g("graph", "euler_faces")),
        ("Face degree CV", g("graph", "face_degree_cv")),
        ("Bigons", g("graph", "n_bigons")),
        ("Plane drawings", g("graph", "n_plane_drawings")),
        ("Symmetry order", g("graph", "automorphism_order")),
        ("Symmetry group", lambda m: m["graph"].get("symmetry_group", "")),
        ("Edge length CV", g("geom2d", "edge_length_cv")),
        ("Dirichlet energy", g("geom2d", "dirichlet_energy_norm")),
        ("Crossing angle dev (deg)", g("geom2d", "crossing_angle_rms_dev_deg")),
        ("2D symmetry score", g("geom2d", "sym2d_score")),
        ("Thomson energy", g("sphere3d", "thomson_energy")),
        ("Sphere spread quality", g("sphere3d", "sphere_spread_quality")),
        ("3D strand length CV", g("sphere3d", "strand3d_length_cv")),
        ("Bending energy", g("sphere3d", "bending_energy")),
        ("3D dot-pattern symmetry", g("sphere3d", "sym3d_order")),
        ("3D point group", lambda m: _point_group_3d(m["dt"])),
        ("q: strand balance", g("quality", "strand_balance")),
        ("q: diagram symmetry", g("quality", "diagram_symmetry")),
        ("q: face regularity", g("quality", "face_regularity")),
        ("q: sphere energy", g("quality", "sphere_energy")),
        ("COMPOSITE", lambda m: m["composite"]),
        ("Linking numbers (sorted)", lambda m: str(m.get("linking_fp"))),
        ("Jones (Sage only)", lambda m: m["jones"]),
    ]


# Columns kept for reference but NOT part of the composite score: the 2-D
# Tutte/geom2d numbers depend on which face draw_dt turns to the outside (the
# 'puncture'), so they are descriptive only and shown on a grey background.
_DESCRIPTIVE_COLS = {
    "Edge length CV",
    "Dirichlet energy",
    "Crossing angle dev (deg)",
    "2D symmetry score",
    "3D dot-pattern symmetry",
}


def write_excel(scored, path, run_info):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cols = _cols()
    wb = Workbook()
    ws = wb.active
    ws.title = "representatives"

    head_fill = PatternFill("solid", fgColor="1F3864")
    desc_head_fill = PatternFill("solid", fgColor="808080")   # descriptive-only header
    desc_fill = PatternFill("solid", fgColor="EDEDED")         # descriptive-only cell
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    best_fill = PatternFill("solid", fgColor="C6EFCE")
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cidx, (header, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=cidx, value=header)
        cell.fill = desc_head_fill if header in _DESCRIPTIVE_COLS else head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for ridx, m in enumerate(scored, start=2):
        best = (m["rank"] == 1)
        for cidx, (header, getter) in enumerate(cols, start=1):
            val = getter(m)
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Descriptive 2-D columns stay grey even on the best row: they are not
            # part of the composite (puncture-dependent, see _DESCRIPTIVE_COLS).
            if header in _DESCRIPTIVE_COLS:
                cell.fill = desc_fill
            elif best:
                cell.fill = best_fill

    # widths
    for cidx, (header, _) in enumerate(cols, start=1):
        letter = get_column_letter(cidx)
        # NB: match the headers _cols() actually emits.  These tested "DT code" and
        # "Jones (same link check)", neither of which exists any more, so the widest
        # column in the sheet (a 77-character canonical DT code) was being squeezed
        # into the 19-wide default.
        if header == "Canonical DT code":
            ws.column_dimensions[letter].width = 42
        elif header in ("Jones (Sage only)", "Strand lengths", "Clasp structure",
                        "Linking numbers (sorted)"):
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = max(11, min(20, len(header) + 2))
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(scored) + 1)

    # run info sheet
    ws2 = wb.create_sheet("run_info")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 80
    for r, (k, v) in enumerate(run_info.items(), start=1):
        a = ws2.cell(row=r, column=1, value=k)
        a.font = Font(name="Arial", bold=True, size=10)
        b = ws2.cell(row=r, column=2, value=str(v))
        b.font = Font(name="Arial", size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")

    # legend sheet: for every metric column, whether higher or lower is better
    ws3 = wb.create_sheet("metric_legend")
    for col, w in (("A", 26), ("B", 20), ("C", 74)):
        ws3.column_dimensions[col].width = w
    green = PatternFill("solid", fgColor="C6EFCE")   # higher = better
    blue = PatternFill("solid", fgColor="DDEBF7")    # lower = better
    grey = PatternFill("solid", fgColor="EDEDED")    # descriptive / fixed
    for c, txt in enumerate(("Metric", "Direction", "Meaning"), start=1):
        cell = ws3.cell(row=1, column=c, value=txt)
        cell.fill = head_fill
        cell.font = head_font
        cell.border = border
    fillmap = {"higher = better": green, "lower = better": blue,
               "descriptive": grey, "fixed": grey}
    for r, (metric, direction, meaning) in enumerate(_METRIC_LEGEND, start=2):
        for c, val in enumerate((metric, direction, meaning), start=1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if c == 2:
                cell.fill = fillmap.get(direction, grey)

    wb.save(path)


# For each Excel metric column: is larger better, smaller better, or just descriptive?
_METRIC_LEGEND = [
    ("Rank", "lower = better", "1 = best blueprint overall."),
    ("Diagram name", "descriptive", "Structural name of the diagram: 'Orthogonal rosette' (clasp-free, C2xC2) or one of the Bing-double clasps (Balanced 4-4, Offset 5-3 a/b, Lopsided 6-2)."),
    ("Clasp structure", "descriptive", "The Bing-double reading: two frame components carry 3 crossings each; the two clasp components split 8 crossings; the split (4-4 / 5-3 / 6-2) sets the clasp balance. The orthogonal form has no localized clasp."),
    ("Multiplicity", "descriptive", "How often the simplifier produced this diagram; NOT a quality (the best diagram is actually rare)."),
    ("First round", "descriptive", "Round at which the diagram was first seen."),
    ("Crossings", "lower = better", "Diagram complexity; fixed at 14 here (all are minimal)."),
    ("Components", "fixed", "Number of strands; fixed at 4."),
    ("Strand length CV", "lower = better", "Spread of strand lengths; 0 = all strands equal length."),
    ("Strand balance entropy", "higher = better", "Evenness of strand lengths; 1 = perfectly even."),
    ("Strand length max/min", "lower = better", "Longest / shortest strand; 1 = equal."),
    ("Self crossings", "descriptive", "Within-strand crossings; 0 for this link."),
    ("Inter crossings", "descriptive", "Between-strand crossings."),
    ("Linking degree CV", "lower = better", "Evenness of how the linking load is shared; lower = more even."),
    ("Alternating fraction", "descriptive", "How close to an alternating diagram; not scored."),
    ("Over/under imbalance", "descriptive", "Convention-dependent over/under balance; not scored."),
    ("Faces (n+2)", "fixed", "Number of regions; equals crossings + 2 (Euler check) = 16."),
    ("Face degree CV", "lower = better", "Regularity of the regions; lower = more uniform tiling."),
    ("Bigons", "lower = better", "Two-sided regions (clasps); 0 = no local crowding."),
    ("Plane drawings", "lower = more symmetric",
     "How many genuinely different plane pictures the diagram has: every face punctured in turn "
     "(bigons included), grouped up to rotation, in-plane mirror, component swap, strand reversal "
     "and the complete crossing flip. INTRINSIC -- unlike the 2-D metrics it does not depend on "
     "which face is turned to the outside. Not scored; reported for comparison. "
     "enumerate_puncturing_dt.py draws the atlas."),
    ("Symmetry order", "higher = better", "Combinatorial symmetry: how many DT re-encodings fix the diagram (RESPECTS over/under); higher = fewer unique sequences."),
    ("Symmetry group", "descriptive", "The symmetry group type respecting over/under (e.g. C1 trivial, C2, C2xC2 = Klein four-group of order 4, Ck cyclic). C2xC2 has order 4 but is NOT a 4-fold rotation."),
    ("Edge length CV", "descriptive", "DESCRIPTIVE ONLY (not scored): uniformity of segment lengths in the relaxed 2-D layout. Depends on which face draw_dt turns to the outside (the 'puncture'), which is not intrinsic when faces tie for largest -- so it is excluded from the composite."),
    ("Dirichlet energy", "descriptive", "DESCRIPTIVE ONLY (not scored): 2-D spring energy. Puncture-dependent (see Edge length CV), so excluded from the composite."),
    ("Crossing angle dev (deg)", "descriptive", "DESCRIPTIVE ONLY (not scored): deviation of crossings from ideal 90-degree X shapes in the 2-D layout. Puncture-dependent, so excluded from the composite."),
    ("2D symmetry score", "descriptive", "DESCRIPTIVE ONLY (not scored): rotational symmetry of the crossing POSITIONS in the 2-D layout; IGNORES over/under AND depends on the puncture, so excluded from the composite. Use 'Symmetry order' (sign-aware) and '3D point group' instead. A score of 0 means no rotational symmetry was found at all (order 1), the same convention as the 3-D column."),
    ("Thomson energy", "lower = better", "Crowding of crossings on the sphere; lower = more evenly spread."),
    ("Sphere spread quality", "higher = better", "Evenness of the crossing spread on the sphere, as a ratio to a Fibonacci-spiral reference of the same number of points. CAPPED AT 1: the reference is evenly spread but not the true minimum-energy arrangement, so a small very even diagram can beat it; the uncapped ratio is kept as sphere_spread_raw in the JSON."),
    ("3D strand length CV", "lower = better", "Evenness of strand lengths measured on the 3-D sphere."),
    ("Bending energy", "lower = better", "How sharply strands turn in 3-D; lower = gentler, relaxed curves."),
    ("3D dot-pattern symmetry", "descriptive", "DESCRIPTIVE ONLY (not scored): largest rotational symmetry of the crossing POSITIONS only in the 3-D layout (a k-fold dot pattern); IGNORES over/under and crossing signs, so it false-positives on low-symmetry diagrams (e.g. a spurious 2-fold on the C1 Lopsided and Cs Offset clasps) -- excluded from the composite. The composite's 'q: diagram symmetry' uses the loop-gated '3D point group' order instead. Use 'Symmetry order'/'Symmetry group'/'3D point group' for the sign-respecting values."),
    ("3D point group", "descriptive", "True Schoenflies point group of the 3-D spherical embedding (up to mirror): rotation axes (Cn) classified vs improper operations sub-typed by eigenvalues -- mirror plane (Cs), inversion centre (Ci), rotoreflection (Sn) -- with a full-loop reliability gate. E.g. Orthogonal rosette C2v, Balanced clasp Ci (inversion centre; NOT Cs), Offset clasp Cs, Lopsided clasp C1."),
    ("q: strand balance", "higher = better", "0-1 quality for even strands."),
    ("q: diagram symmetry", "higher = better", "0-1 quality for symmetry: mean of the sign-aware combinatorial automorphism order and the loop-gated 3-D point-group order (both as 1 - 1/order). Both are topological; the dot-only 3-D dot-pattern symmetry and the puncture-dependent 2-D symmetry are descriptive only and NOT included."),
    ("q: face regularity", "higher = better", "0-1 quality for a regular planar tiling: low face-degree CV and few bigons. Computed from the planar embedding (independent of the puncture); replaces the former 2-D 'geometric strain'."),
    ("q: sphere energy", "higher = better", "0-1 quality for even, relaxed 3-D layout."),
    ("COMPOSITE", "higher = better", "Overall 0-1 score; higher = better synthesis blueprint."),
    ("Linking numbers (sorted)", "descriptive", "Pairwise linking fingerprint (same for all: the link is fixed)."),
    ("Jones (Sage only)", "descriptive", "Link invariant; identical for all (same link). Populates under Sage."),
]


# --------------------------------------------------------------------------- #
#  4b. SVG figure  (2-D Tutte layout + 3-D sphere layout per representative)
# --------------------------------------------------------------------------- #
_PALETTE = ["#4c72b0", "#dd8452", "#55a868", "#c44e52", "#8172b3",
            "#937860", "#da8bc3", "#8c8c8c", "#ccb974", "#64b5cd"]


def _arc_list(model):
    """Directed traversal arcs (ci, a, b) between crossing indices, one per visit-step."""
    pos_cross = model["pos_cross"]
    arcs = []
    for ci, cp in enumerate(model["comp_positions"]):
        ks = [pos_cross[p] for p in cp]
        L = len(ks)
        for i in range(L):
            arcs.append((ci, ks[i], ks[(i + 1) % L]))
    return arcs


def _bow_offsets(arcs):
    """Offset multiplier for each arc so that arcs sharing the same endpoint pair
    (e.g. the two arcs of a bigon, or arcs of different components between the same
    crossings) fan apart instead of drawing on top of each other.  Single edges get 0."""
    from collections import defaultdict
    groups = defaultdict(list)
    for idx, (ci, a, b) in enumerate(arcs):
        groups[tuple(sorted((a, b)))].append(idx)
    off = {}
    for _, idxs in groups.items():
        k = len(idxs)
        for j, idx in enumerate(idxs):
            off[idx] = (j - (k - 1) / 2.0)     # symmetric around 0
    return off


def _draw_skeleton_2d(ax, C2, model, gap_frac=0.05):
    """2-D component-coloured skeleton; parallel arcs are bowed apart (quadratic Bezier)."""
    arcs = _arc_list(model)
    off = _bow_offsets(arcs)
    span = float(np.max(C2.max(0) - C2.min(0))) or 1.0
    gap = gap_frac * span
    ts = np.linspace(0, 1, 26)[:, None]
    for idx, (ci, a, b) in enumerate(arcs):
        pa, pb = C2[a], C2[b]
        col = _PALETTE[ci % len(_PALETTE)]
        # Use a CANONICAL direction (sorted endpoints) for the perpendicular so that two
        # parallel arcs between the same crossings bow to OPPOSITE sides regardless of the
        # direction each is traversed.  (Using each arc's own direction flips the perp and
        # made oppositely-traversed bigon arcs overlap.)
        lo, hi = (a, b) if a <= b else (b, a)
        dc = C2[hi] - C2[lo]
        L = float(np.hypot(dc[0], dc[1]))
        if L < 1e-9:
            continue
        perp = np.array([-dc[1], dc[0]]) / L
        ctrl = (pa + pb) / 2.0 + perp * (off[idx] * gap * 2.0)   # midpoint bow = off*gap
        pts = (1 - ts) ** 2 * pa + 2 * (1 - ts) * ts * ctrl + ts ** 2 * pb
        ax.plot(pts[:, 0], pts[:, 1], "-", lw=1.5, color=col, alpha=0.9)
    ax.plot(C2[:, 0], C2[:, 1], "o", ms=4.0, color="#222222", zorder=3)
    ax.set_aspect("equal")
    ax.axis("off")


def _perm_order_local(p):
    seen, order = set(), 1
    for start in range(len(p)):
        if start in seen:
            continue
        L, x = 0, start
        while True:
            seen.add(x); x = p[x]; L += 1
            if x == start:
                break
        order = order * L // math.gcd(order, L)
    return order


_SYM3D_CACHE = {}


def _compute_sym3d(dt, C3):
    """Run the V2_0 rotation-system + eigenvalue engine once for `dt` in the frame
    of C3 and cache both the typed elements and the short point-group label
    (Cn / Cs / Ci / Sn ...).  A full-loop reliability gate is applied inside."""
    if _too_complex(dt, "3-D point group"):
        # CDT2.symmetry_3d enumerates the same relabellings (_flipfalse_syms), so it hangs
        # on exactly the diagrams _mirror_canonical does.  order 0 makes
        # _point_group_order_3d fall back to the automorphism order.
        out = {"order": 0, "elements": [], "point_group": "n/a (too complex)"}
        _SYM3D_CACHE[dt] = out
        return out
    try:
        comps = tuple(tuple(c) for c in CDT2.parse_dt(dt))          # dt is already canonical here
        C3 = np.asarray(C3, float)
        res = CDT2.symmetry_3d(comps, centers=C3, loops=CDT2._weave_loops(dt, C3))
        out = {"order": int(res.get("order", 0)),                   # loop-gated point-group order (incl. identity)
               "elements": list(res.get("elements", [])),
               "point_group": res.get("point_group", "").split("(")[0].strip()}
    except Exception:  # noqa: BLE001
        out = {"order": 0, "elements": [], "point_group": ""}
    _SYM3D_CACHE[dt] = out
    return out


def _symmetry_axes_3d(dt, C3):
    """True point-group elements (list of (kind, order, vec); kind in
    {'axis','mirror','inversion','improper-axis'}, vec None for inversion), in the
    frame of C3.  Replaces the old det-only method that mislabelled every improper
    op 'mirror' (e.g. the Balanced clasp's inversion centre)."""
    if dt in _SYM3D_CACHE:
        return _SYM3D_CACHE[dt]["elements"]
    return _compute_sym3d(dt, C3)["elements"]


def _point_group_3d(dt):
    """Short Schoenflies label of the diagram's true 3-D point group (from the
    shared cache; populated once per representative during scoring)."""
    return _SYM3D_CACHE.get(dt, {}).get("point_group", "")


def _point_group_order_3d(dt, C3, aut_fallback=1):
    """Order of the diagram's loop-aware 3-D point group -- the number of symmetry
    operations including the identity (C1 -> 1, C2/Cs/Ci -> 2, C2v -> 4, D3 -> 6).

    Uses the same loop-gated engine as the '3D point group' label
    (CDT2.symmetry_3d), so it counts only operations that map the actual 3-D loops
    -- strands woven in/out with over/under -- onto themselves.  This is the
    topological symmetry, unlike _best_rotational_symmetry_3d, which inspects only
    the crossing-dot positions and false-positives on low-symmetry diagrams (it
    reports a spurious order-2 rotation for the C1 Lopsided and Cs Offset clasps).
    Falls back to the sign-aware combinatorial automorphism order if the 3-D engine
    is unavailable or returns nothing."""
    info = _SYM3D_CACHE.get(dt) or _compute_sym3d(dt, C3)
    order = info.get("order", 0)
    if order and order >= 1:
        return int(order)
    return max(1, int(aut_fallback))


def _draw_symmetry_3d(ax3, sym_elems):
    """Overlay the true point-group elements:
      * rotation axis (Cn)     -> thin short-dashed line + Cn label, extended past the sphere;
      * mirror plane (sigma)   -> translucent square reaching the sphere, thin short-dashed outline;
      * inversion centre (i)   -> a small dot at the centre labelled 'i';
      * rotoreflection (Sn)    -> thin short-dashed axis + Sn label (improper axis)."""
    if not sym_elems:
        return
    from mpl_toolkits.mplot3d.art3d import Poly3DCollection
    dash = (0, (2.4, 1.8))                        # short dashes, small gaps
    # High z-orders + clip_on=False so the axis/label/dot/plane sit ABOVE the faint
    # sphere skeleton (which is drawn at zorder 0) and are never clipped by the panel.
    for kind, order, vec in sym_elems:
        if kind in ("axis", "improper-axis"):
            if vec is None:
                continue
            vec = np.asarray(vec, float); vec = vec / (np.linalg.norm(vec) + 1e-12)
            p0, p1 = -1.6 * vec, 1.6 * vec        # extend out of the sphere
            ax3.plot([p0[0], p1[0]], [p0[1], p1[1]], [p0[2], p1[2]],
                     color="#111111", lw=1.0, ls=dash, alpha=0.95, zorder=30,
                     clip_on=False)
            lbl = ("C%d" % order) if kind == "axis" else ("S%d" % order)
            for s in (1.0, -1.0):                 # label BOTH ends so one is always front-facing
                ax3.text(s * 1.74 * vec[0], s * 1.74 * vec[1], s * 1.74 * vec[2], lbl,
                         color="#111111", fontsize=9, fontweight="bold",
                         ha="center", va="center", zorder=40, clip_on=False)
        elif kind == "inversion":
            # inversion centre: a dot at the origin, labelled i
            ax3.scatter([0.0], [0.0], [0.0], c="#7e3f98", s=46, marker="o",
                        edgecolors="#3a1d47", linewidths=0.6, zorder=32,
                        depthshade=False, clip_on=False)
            ax3.text(0.0, 0.0, 0.18, "i", color="#5b2d70", fontsize=11, fontweight="bold",
                     ha="center", va="center", zorder=40, clip_on=False)
        else:  # mirror plane -> translucent square reaching the sphere, labelled sigma
            if vec is None:
                continue
            n = np.asarray(vec, float); n = n / (np.linalg.norm(n) + 1e-12)
            e1 = np.cross(n, np.array([0.0, 0.0, 1.0]))
            if np.linalg.norm(e1) < 1e-6:
                e1 = np.cross(n, np.array([1.0, 0.0, 0.0]))
            e1 = e1 / np.linalg.norm(e1)
            e2 = np.cross(n, e1); e2 = e2 / (np.linalg.norm(e2) + 1e-12)
            h = 1.05                              # half-side: edges reach the unit sphere
            corners = np.array([h * e1 + h * e2, -h * e1 + h * e2,
                                -h * e1 - h * e2, h * e1 - h * e2])
            poly = Poly3DCollection([corners], facecolor="#7e3f98", edgecolor="none",
                                    alpha=0.06, zorder=1)
            poly.set_clip_on(False)
            ax3.add_collection3d(poly)
            loop = np.vstack([corners, corners[0]])
            ax3.plot(loop[:, 0], loop[:, 1], loop[:, 2],
                     color="#7e3f98", lw=1.0, ls=dash, alpha=0.95, zorder=28,
                     clip_on=False)
            lp = 1.22 * e1                        # sigma label just outside an edge of the square
            ax3.text(lp[0], lp[1], lp[2], "σ", color="#5b2d70", fontsize=10,
                     fontstyle="italic", fontweight="bold", ha="center", va="center",
                     zorder=40, clip_on=False)


def _draw_sphere_depth(ax3, C3, model, elev=22.0, azim=-58.0, gap=0.13, zoom=1.9,
                       sym_elems=None):
    """3-D sphere layout: arcs ride on the sphere (renormalized), parallel arcs are
    bowed apart, and transparency is depth-cued (nearer = opaque, farther = faint)."""
    ax3.view_init(elev=elev, azim=azim)
    er, ar = math.radians(elev), math.radians(azim)
    eye = np.array([math.cos(er) * math.cos(ar),
                    math.cos(er) * math.sin(ar), math.sin(er)], float)
    u = np.linspace(0, 2 * np.pi, 26)
    v = np.linspace(0, np.pi, 13)
    wf = ax3.plot_wireframe(np.outer(np.cos(u), np.sin(v)), np.outer(np.sin(u), np.sin(v)),
                            np.outer(np.ones_like(u), np.cos(v)), color="0.9", linewidth=0.2)
    wf.set_zorder(0); wf.set_clip_on(False)       # skeleton on the lowest layer, never clips
    arcs = _arc_list(model)
    off = _bow_offsets(arcs)
    ss = np.linspace(0, 1, 16)
    for idx, (ci, a, b) in enumerate(arcs):
        pa, pb = C3[a], C3[b]
        col = _PALETTE[ci % len(_PALETTE)]
        # canonical direction (sorted endpoints) so parallel arcs bow to opposite sides
        lo, hi = (a, b) if a <= b else (b, a)
        chord = C3[hi] - C3[lo]
        mid = (pa + pb) / 2.0
        t = np.cross(mid, chord)
        nt = np.linalg.norm(t)
        if nt < 1e-9:
            continue
        t = t / nt
        ctrl = mid + t * (off[idx] * gap * 2.0)
        # bowed arc, each sample renormalized to the unit sphere
        pts = ((1 - ss) ** 2)[:, None] * pa + (2 * (1 - ss) * ss)[:, None] * ctrl \
            + (ss ** 2)[:, None] * pb
        pts = pts / np.linalg.norm(pts, axis=1, keepdims=True)
        for j in range(len(ss) - 1):
            s0, s1 = pts[j], pts[j + 1]
            depth = float(np.dot((s0 + s1) / 2.0, eye))
            alpha = 0.12 + 0.83 * (depth + 1.0) / 2.0
            ax3.plot([s0[0], s1[0]], [s0[1], s1[1]], [s0[2], s1[2]],
                     "-", lw=1.9, color=col, alpha=alpha, solid_capstyle="round",
                     zorder=5, clip_on=False)
    dvals = C3 @ eye
    alphas = 0.2 + 0.8 * (dvals - dvals.min()) / (np.ptp(dvals) + 1e-9)
    for k in range(len(C3)):
        ax3.scatter(C3[k, 0], C3[k, 1], C3[k, 2], c="#222222",
                    s=9, alpha=float(alphas[k]), depthshade=False,
                    zorder=6, clip_on=False)
    _draw_symmetry_3d(ax3, sym_elems)                # overlay C2 axes / mirror planes / i dot
    # EQUAL, symmetric data limits so the sphere renders ROUND.  Without this,
    # Matplotlib auto-scales each axis independently to the plotted data; the symmetry
    # overlays (axis to +-1.6*vec along a diagonal, labels at 1.74, mirror squares to
    # ~1.48) extend the ranges ASYMMETRICALLY, so the cubic box aspect squeezes/stretches
    # the unit sphere.  A fixed cube of half-width 1.8 contains every overlay and keeps
    # all panels the same size.
    _R = 1.8
    ax3.set_xlim(-_R, _R); ax3.set_ylim(-_R, _R); ax3.set_zlim(-_R, _R)
    try:
        ax3.set_box_aspect((1, 1, 1), zoom=zoom)     # zoom fills the panel (bigger sphere)
    except TypeError:
        ax3.set_box_aspect((1, 1, 1))
    ax3.axis("off")


# 2-D layout used for BOTH the draw_dt_original_labels panel and the Tutte skeleton,
# so their orientation/rotation is identical (per request: shaped-tutte, ellipse, aspect 1).
_DRAW_LAYOUT = "shaped-tutte"
_DRAW_TUTTE_OPTS = {"shape": "ellipse", "aspect": 1.0}
_DRAW_MIN_SEP = 0.02          # push apart non-incident strand pieces closer than this (fraction of span)
_SPHERE_VIEWS = [(22.0, -58.0), (22.0, 122.0)]   # two viewpoints (elev, azim), ~180 deg apart


def _render_draw(ax, model, P, centers_d, col_of, show_labels):
    try:
        DDOL.render_diagram(ax, model, P, centers_d, color_of=col_of,
                            show_labels=show_labels, arrows=True, lw=1.7, label_fontsize=5.5)
    except Exception as exc:  # keep the grid robust
        ax.text(0.5, 0.5, "render error:\n%s" % exc, ha="center", va="center",
                fontsize=6, transform=ax.transAxes)
    ax.set_aspect("equal")
    ax.axis("off")


def _strip_clip_paths(fig):
    """Turn off clipping on every artist so the saved SVG contains no <clipPath>
    masks (and nothing is cut off at panel edges).  Iterates all axes and their
    descendant artists, clearing both the clip flag and any assigned clip path."""
    for ax in fig.get_axes():
        try:
            ax.patch.set_clip_on(False)
        except Exception:  # noqa: BLE001
            pass
        for art in ax.findobj():
            try:
                art.set_clip_on(False)
            except Exception:  # noqa: BLE001
                pass
            try:
                art.set_clip_path(None)
            except Exception:  # noqa: BLE001
                pass


def make_figure(scored, path, max_draw=60):
    import textwrap
    reps = scored[:max_draw]
    n = len(reps)
    total = sum(m["_class"]["multiplicity"] for m in scored)
    nsph = len(_SPHERE_VIEWS)
    # columns: text | draw(labelled) | draw(no labels) | skeleton | sphere views...
    ncol = 4 + nsph
    fig = plt.figure(figsize=(3.15 * ncol + 3.6, 3.9 * n + 0.7))
    # wider sphere columns (1.7) so the 3-D drawings are larger; clip is turned off
    # everywhere (see end of this function) so nothing is cut at the panel edges.
    gs = fig.add_gridspec(n, ncol, width_ratios=[1.05, 1.15, 1.15, 0.95] + [1.7] * nsph,
                          hspace=0.30, wspace=0.02)

    for i, m in enumerate(reps):
        model = DDOL.build_model(DDOL.parse_dt(m["dt"]))
        G = DDOL.build_gadget_graph(model)
        # shared shaped-tutte layout -> identical rotation in the drawings and the skeleton
        P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=dict(_DRAW_TUTTE_OPTS))
        if _DRAW_MIN_SEP > 0:
            P = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP)
        centers_d = DDOL.crossing_centers(model, P)
        C2 = np.array([centers_d[k] for k in range(len(model["crossings"]))], float)
        C3 = m["sphere3d"]["_centers3d"]
        rid = m["_class"]["rep_id"]
        col_of = lambda ci: _PALETTE[ci % len(_PALETTE)]
        group = m["graph"].get("symmetry_group", "C%d" % m["graph"]["automorphism_order"])

        # --- col 0: text (full CANONICAL DT code + metrics) ---
        axt = fig.add_subplot(gs[i, 0])
        axt.axis("off")
        name, clasp = _diagram_name(m["dt"])
        name_line = ("%s\n" % name) if name else ""
        clasp_line = ("%s\n" % "\n".join(textwrap.wrap(clasp, width=34))) if clasp else ""
        head = ("Rank %d   (rep #%d)\n%s%scomposite = %.3f\nmultiplicity = %d / %d\n"
                "symmetry group = %s (order %d)\n3D dot pattern = %d-fold\n"
                "strands = %s\nbending = %.1f"
                % (m["rank"], rid, name_line, clasp_line,
                   m["composite"], m["_class"]["multiplicity"], total,
                   group, m["graph"]["automorphism_order"], m["sphere3d"]["sym3d_order"],
                   m["combinatorial"]["strand_visit_lengths"],
                   m["sphere3d"]["bending_energy"]))
        dt_wrapped = "\n".join(textwrap.wrap(m["dt"], width=30))
        axt.text(0.0, 1.0, head, transform=axt.transAxes, ha="left", va="top",
                 fontsize=9.5, fontweight="bold" if m["rank"] == 1 else "normal")
        axt.text(0.0, 0.30, "Canonical DT code:", transform=axt.transAxes, ha="left", va="top",
                 fontsize=8.5, style="italic")
        axt.text(0.0, 0.23, dt_wrapped, transform=axt.transAxes, ha="left", va="top",
                 fontsize=8.0, family="monospace")

        # --- col 1: draw_dt_original_labels WITH DT labels ---
        axd = fig.add_subplot(gs[i, 1])
        _render_draw(axd, model, P, centers_d, col_of, show_labels=True)
        if i == 0:
            axd.set_title("draw_dt_original_labels\n(with DT labels)", fontsize=9)

        # --- col 2: draw_dt_original_labels WITHOUT labels (strands unobscured) ---
        axd2 = fig.add_subplot(gs[i, 2])
        _render_draw(axd2, model, P, centers_d, col_of, show_labels=False)
        if i == 0:
            axd2.set_title("draw_dt_original_labels\n(no labels)", fontsize=9)

        # --- col 3: 2-D skeleton (same layout, parallel arcs bowed apart) ---
        ax2 = fig.add_subplot(gs[i, 3])
        _draw_skeleton_2d(ax2, C2, model)
        if i == 0:
            ax2.set_title("2-D skeleton\n(same layout)", fontsize=9)

        # --- cols 4+: 3-D sphere, two viewpoints, depth-shaded, enlarged ---
        # symmetry elements (C2 axes / mirror planes) overlaid on the 3-D layout
        sym_elems = _symmetry_axes_3d(m["dt"], C3)
        n_ax = sum(1 for k, o, v in sym_elems if k == "axis")
        n_mir = sum(1 for k, o, v in sym_elems if k == "mirror")
        n_inv = sum(1 for k, o, v in sym_elems if k == "inversion")
        n_S = sum(1 for k, o, v in sym_elems if k == "improper-axis")
        parts = []
        if n_ax:
            parts.append("C%d axis" % max((o for k, o, v in sym_elems if k == "axis"), default=2))
        if n_mir:
            parts.append("%d mirror%s" % (n_mir, "" if n_mir == 1 else "s"))
        if n_inv:
            parts.append("inversion centre (i)")
        if n_S:
            parts.append("%d S%d" % (n_S, max((o for k, o, v in sym_elems if k == "improper-axis"), default=2)))
        sym_label = "3-D symmetry: " + (", ".join(parts) if parts else "none (C1)")
        for vi, (elev, azim) in enumerate(_SPHERE_VIEWS):
            ax3 = fig.add_subplot(gs[i, 4 + vi], projection="3d")
            _draw_sphere_depth(ax3, C3, model, elev=elev, azim=azim, sym_elems=sym_elems)
            if i == 0:
                ax3.set_title("3-D sphere (view %d)\n%s" % (vi + 1, sym_label), fontsize=8)
            elif vi == 0:
                ax3.set_title(sym_label, fontsize=7)

    fig.suptitle("Representative diagrams of one link, ranked by composite score  "
                 "(canonical DT codes at left; 3-D sphere from two viewpoints)",
                 fontsize=13, y=0.999)
    fig.text(0.5, 0.004,
             "On the 3-D spheres (true point group, up to mirror): a black dashed line = a "
             "rotation (Cn) axis, a violet square labelled σ = a mirror plane, a violet dot "
             "labelled i at the centre = an inversion centre, a dashed axis + Sn = a rotoreflection.",
             ha="center", va="bottom", fontsize=8, color="#555555")
    _strip_clip_paths(fig)                   # remove all clipping masks from the SVG
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def _draw_tutte_opts(puncture=None):
    """Copy of the shared draw options, optionally pinning a specific outer
    ('puncture') face by its crossing-ID signature (a tuple of 'cN' tokens)."""
    opts = dict(_DRAW_TUTTE_OPTS)
    if puncture is not None:
        opts["puncture_face"] = puncture
    return opts


def _draw_one(ax, dt, col_of, show_labels=False, rasterize=False, puncture=None):
    """Render a single DT code with draw_dt_original_labels in its own shaped-tutte
    layout.  ``puncture`` pins which face is turned to the outside, so the SAME
    diagram can be drawn with different outer faces: pass a face KEY (a frozenset
    of graph nodes, from _diagram_face_records) to pin a face exactly, or a
    crossing-ID signature tuple for the old signature-based selection."""
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    if isinstance(puncture, frozenset):
        _, P = _positions_for_face(model, G, puncture)
    else:
        P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=_draw_tutte_opts(puncture))
        if _DRAW_MIN_SEP > 0:
            P = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP)
    centers_d = DDOL.crossing_centers(model, P)
    _render_draw(ax, model, P, centers_d, col_of, show_labels)
    if rasterize:
        ax.set_rasterized(True)      # embed as a small raster in SVG -> fast, valid SVG


def _is_seg_node(node):
    return isinstance(node, tuple) and len(node) == 2 and node[0] == "seg"


def _diagram_face_records(model, emb, crossing_ids=None):
    """Every face of the DIAGRAM, as records carrying the face's identity.

    The gadget graph turns each crossing into a 4-cycle of corner nodes, and that
    4-cycle is a face of the graph but not of the diagram; it carries no
    ('seg', p) midpoint node, which is how it is recognised and dropped.  For a
    real face the number of segment nodes IS its polygon degree, so a bigon has
    two.

    ``key`` is the frozenset of the face's nodes -- its identity.  A face must be
    pinned by identity and not by its crossing-ID signature, because two
    different faces can touch the same crossings (the two triangles of the
    standard trefoil do) and a signature cannot separate those.
    """
    if crossing_ids is None:
        crossing_ids = DDOL.default_crossing_ids(model)
    recs = []
    for face in DDOL.planar_faces(emb):
        segs = sorted(int(v[1]) for v in face if _is_seg_node(v))
        if not segs:
            continue                       # crossing-gadget interior, not a face
        recs.append({
            "key": frozenset(face),
            "sig": DDOL._face_signature(face, crossing_ids),
            "n_edges": len(segs),
            "crossings": DDOL._face_crossing_indices(face),
            "edges": tuple(segs),
        })
    recs.sort(key=lambda r: (-r["n_edges"], tuple(r["crossings"]), r["edges"]))
    return recs


def _marked_graph(model, G, face_key, flip=False):
    """The gadget graph, node-labelled by over/under and by the punctured face.

    Two such graphs are isomorphic exactly when a symmetry of the diagram carries
    one puncture to the other -- so isomorphism IS the question "do these two
    punctures give the same drawing", answered exactly and without laying
    anything out.

    Component identity and the in_*/out_* roles are deliberately not labelled, so
    the test is up to component swap and strand-orientation reversal; the graph
    being undirected makes rotations and in-plane mirrors free.  ``flip`` swaps
    every 'O' with every 'U': that is the mirror THROUGH the plane of the paper,
    which leaves the shadow alone and reverses every crossing at once -- the
    ordinary mirror image, and the same up-to-mirror convention `dedup` already
    applies at the class level via _mirror_canonical.
    """
    H = G.copy()
    over_at = model["over_at"]
    hi, lo = ("U", "O") if flip else ("O", "U")
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        over_o = bool(over_at[cr["odd"]])
        over_e = bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = hi if over_o else lo
        for role in ("in_e", "out_e"):
            labels[(k, role)] = hi if over_e else lo
    for node in H.nodes():
        labels[node] = labels.get(node, "S") + ("*" if node in face_key else "")
    nx.set_node_attributes(H, labels, "lab")
    return H


def _puncture_orbits(model, G, recs, chiral_merge=True):
    """Group punctured faces into orbits of the diagram's symmetry group.

    Buckets on a Weisfeiler-Lehman hash, then confirms with exact VF2.  Replaces
    the old geometric grouping, which compared only the rendered crossing
    POSITIONS: the Tutte layout depends solely on the shadow, so two punctures
    giving congruent skeletons that differ in which strand goes over produced the
    same key and were wrongly merged -- the figure showed one panel fewer than
    exists for three of the four project diagrams.  This test is exact, needs no
    layout at all, and is independent of the drawing settings.
    """
    nm = lambda a, b: a["lab"] == b["lab"]  # noqa: E731
    described = []
    for r in recs:
        forms = [_marked_graph(model, G, r["key"])]
        if chiral_merge:
            forms.append(_marked_graph(model, G, r["key"], flip=True))
        hashes = [nx.weisfeiler_lehman_graph_hash(H, node_attr="lab", iterations=5)
                  for H in forms]
        described.append((r, forms, min(hashes)))
    buckets, order = {}, []
    for r, forms, h in described:
        if h not in buckets:
            buckets[h] = []
            order.append(h)
        buckets[h].append((r, forms))
    orbits = []
    for h in order:
        items = buckets[h]
        if len(items) == 1:
            orbits.append([items[0][0]])
            continue
        groups = []
        for r, forms in items:
            for rep_H, members in groups:
                if any(nx.is_isomorphic(H, rep_H, node_match=nm) for H in forms):
                    members.append(r)
                    break
            else:
                groups.append((forms[0], [r]))
        orbits.extend(members for _, members in groups)
    return orbits


def _pinned_outer_face(face_key, original):
    """A stand-in for DDOL.select_outer_face returning one specific face.

    The public selector takes a crossing-ID signature, which does not always
    identify a face uniquely; this pins by node identity instead.  Installed only
    around a single compute_positions call, and always restored.
    """
    def _select(faces, crossing_ids=None, prefer=None, report_out=None, layout_name=""):
        for f in faces:
            if frozenset(f) == face_key:
                return f
        return original(faces, crossing_ids=crossing_ids, prefer=prefer,
                        report_out=report_out, layout_name=layout_name)
    return _select


def _positions_for_face(model, G, face_key=None):
    """(exact Tutte solve, drawing positions) with ``face_key`` turned outward."""
    opts = dict(_DRAW_TUTTE_OPTS)
    if face_key is None:
        P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=opts)
    else:
        original = DDOL.select_outer_face
        DDOL.select_outer_face = _pinned_outer_face(face_key, original)
        try:
            P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=opts)
        finally:
            DDOL.select_outer_face = original
    drawn = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP) if _DRAW_MIN_SEP > 0 else P
    return P, drawn


def _default_outer_face_key(model, emb, crossing_ids=None):
    """Identity of the face draw_dt punctures on its own (no puncture pinned)."""
    if crossing_ids is None:
        crossing_ids = DDOL.default_crossing_ids(model)
    faces = DDOL.planar_faces(emb)
    outer = DDOL.select_outer_face(faces, crossing_ids=crossing_ids, prefer=None)
    if outer is None or not any(_is_seg_node(v) for v in outer):
        return None
    return frozenset(outer)


def _largest_tie_face_signatures(dt):
    """Crossing-ID signatures of the faces TIED for the largest boundary of a
    diagram's planar embedding -- exactly the faces draw_dt could turn to the
    outside (the 'puncture' choices).  When several tie, each gives a genuinely
    different plane drawing of the SAME diagram; when one face is uniquely
    largest there is a single (canonical) drawing.  Signatures are tuples of
    'cN' tokens, matching draw_dt's puncture-face selector."""
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    ok, emb = nx.check_planarity(G)
    if not ok:
        return []
    faces = DDOL.planar_faces(emb)
    if not faces:
        return []
    crossing_ids = DDOL.default_crossing_ids(model)
    mx = max(len(f) for f in faces)
    sigs, seen = [], set()
    for f in faces:
        if len(f) != mx:
            continue
        sig = DDOL._face_signature(f, crossing_ids)
        if sig not in seen:
            seen.add(sig)
            sigs.append(sig)
    return sigs


def _puncture_distinct_drawings(dt, want, exclude=None, chiral_merge=True):
    """The genuinely different plane drawings this diagram's largest faces give.

    Returns up to ``want`` orbit dicts ``{"rep", "members"}`` of face records --
    one per distinct drawing among the faces tied for the largest boundary,
    which are exactly the faces draw_dt could turn to the outside.  The orbit
    holding the default (unpinned) puncture is dropped, since the canonical
    panel shows it already.

    Grouped by exact marked-graph isomorphism, not by comparing rendered
    positions: the old geometric key could not see over/under and merged
    genuinely different drawings.  ``chiral_merge`` also identifies a drawing
    with its through-the-paper mirror, matching dedup's up-to-mirror classes.
    """
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    ok, emb = nx.check_planarity(G)
    if not ok:
        return []
    recs = _diagram_face_records(model, emb)
    if not recs:
        return []
    mx = max(r["n_edges"] for r in recs)
    tied = [r for r in recs if r["n_edges"] == mx]
    skip = set(exclude or ())
    default_key = _default_outer_face_key(model, emb)
    out = []
    for members in _puncture_orbits(model, G, tied, chiral_merge=chiral_merge):
        if default_key is not None and any(m["key"] == default_key for m in members):
            continue                       # the canonical panel already shows it
        if members[0]["key"] in skip:
            continue
        out.append({"rep": members[0], "members": members})
        if len(out) >= want:
            break
    return out


def _crossing_distance_key(model, P, ndigits=3):
    """Sorted, scale-normalised multiset of pairwise crossing distances.

    Unchanged by rotating, mirroring or translating the picture (and by strand
    reversal, which leaves the crossings put), so it identifies a drawing's
    SKELETON.  It says nothing about over/under -- the Tutte layout depends only
    on the shadow -- which is why the panel grouping uses marked-graph
    isomorphism instead and this is kept only for the equivariance check.
    """
    from itertools import combinations
    C = DDOL.crossing_centers(model, P)
    pts = np.array([C[k] for k in range(len(model["crossings"]))], float)
    if len(pts) < 2:
        return (len(pts),)
    pts = pts - pts.mean(axis=0)
    d = np.sort(np.array([np.hypot(*(pts[i] - pts[j]))
                          for i, j in combinations(range(len(pts)), 2)]))
    mx = d[-1] if d[-1] > 0 else 1.0
    return tuple(np.round(d / mx, ndigits))


def _panel_picture(model, P):
    """Crossing positions plus the direction of the strand passing OVER at each --
    the geometry that decides whether two panels are the same PICTURE.  Both parts
    transform covariantly under rotation, mirroring and scaling."""
    C = DDOL.crossing_centers(model, P)
    over_at = model["over_at"]
    pts, dirs = [], []
    for k, cr in enumerate(model["crossings"]):
        role = "o" if over_at[cr["odd"]] else "e"
        v = np.asarray(P[(k, "out_" + role)], float) - np.asarray(P[(k, "in_" + role)], float)
        pts.append(C[k])
        dirs.append(v / (np.linalg.norm(v) or 1.0))
    return np.array(pts, float), np.array(dirs, float)


def _alignment_mismatches(A, B, tol=2e-3):
    """Over EVERY rigid alignment of two panels' skeletons, how many crossings have
    their over/under swapped; [] when the skeletons are not congruent.

    The whole list matters, not its minimum: a pair can admit one alignment that
    disagrees at a few crossings AND another that disagrees at every one.  The
    latter is the mirror through the plane of the paper, which makes the two the
    same drawing up to chirality -- reporting only the closest alignment hides
    exactly that case.
    """
    (pa, va), (pb, vb) = A, B
    if len(pa) != len(pb) or len(pa) < 2:
        return []

    def _unit(pts):
        q = pts - pts.mean(axis=0)
        return q / (np.sqrt((q ** 2).sum(axis=1)).max() or 1.0)

    qa, qb = _unit(pa), _unit(pb)
    n = len(qa)
    i0 = int(np.argmax(np.linalg.norm(qa, axis=1)))
    counts = []
    for mirror in (False, True):
        flipx = np.array([1.0, -1.0]) if mirror else np.array([1.0, 1.0])
        Ma, Va = qa * flipx, va * flipx
        ra = np.linalg.norm(Ma[i0])
        ang = np.arctan2(Ma[i0][1], Ma[i0][0])
        for j0 in range(n):
            if abs(np.linalg.norm(qb[j0]) - ra) > tol:
                continue
            th = np.arctan2(qb[j0][1], qb[j0][0]) - ang
            R = np.array([[np.cos(th), -np.sin(th)], [np.sin(th), np.cos(th)]])
            RA, RV = Ma @ R.T, Va @ R.T
            sigma, used, ok = {}, set(), True
            for i in range(n):
                d = np.linalg.norm(qb - RA[i], axis=1)
                j = int(np.argmin(d))
                if d[j] > tol or j in used:
                    ok = False
                    break
                sigma[i] = j
                used.add(j)
            if not ok:
                continue
            counts.append(sum(1 for i in range(n)
                              if abs(abs(float(RV[i] @ vb[sigma[i]])) - 1.0) > 1e-2))
    return sorted(counts)


def _draw_congruence_key(dt, ndigits=3, puncture=None):
    """Signature of a diagram's DRAWING (its draw_dt layout), invariant under planar
    rotation, reflection (flip) and strand-direction reversal -- i.e. exactly the rigid
    moves of the picture, and nothing else.  It is computed ONLY from the rendered
    crossing positions (no DT-canonical form, no graph signature): the sorted, scale-
    normalised multiset of pairwise crossing distances, which is unchanged by rotating,
    mirroring or translating the picture (and by strand reversal, which leaves crossing
    positions put).  Two raw codes share this key iff draw_dt draws them as the same
    picture up to those moves; two codes of the SAME group that draw differently -- e.g.
    with a different face turned to the outside -- get DIFFERENT keys and are both kept.

    Deliberately computed on the RAW Tutte solve, before ``nudge_min_separation``.
    The nudge is an iterative, order-dependent cosmetic repair, and it perturbs
    two identical drawings by different amounts: on the Balanced clasp it moves
    the two bigon-punctured drawings apart by 1.3% of the diagram span, which is
    far more than this key's rounding, so one picture was counted as two.  Without
    it those two agree to machine precision, as the diagram's inversion symmetry
    says they must.  The panels are still DRAWN with the nudge (see _draw_one) --
    it is a rendering step, not part of what makes two pictures the same."""
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=_draw_tutte_opts(puncture))
    return _crossing_distance_key(model, P, ndigits)


def make_raw_grouping_figure(classes, path, max_per_class=6, max_classes=40,
                             rasterize=False, verify=True):
    """Show how the many RAW sampled diagrams collapse into the few de-duplicated ones.
    Each row is one class: a text summary, its canonical diagram, then several RAW member
    codes.  The raw members are drawn in their own layout, so they look like rotations /
    reflections of each other -- which is exactly the equivalence used to group them."""
    import textwrap
    reps = classes[:max_classes]
    K = max_per_class
    ncol = 2 + K                                  # text | canonical | puncture drawings...
    nrows = len(reps)
    fig = plt.figure(figsize=(2.7 * ncol + 2.0, 3.25 * nrows + 1.0))
    gs = fig.add_gridspec(nrows, ncol, width_ratios=[1.5, 1.25] + [1.1] * K,
                          hspace=0.62, wspace=0.04)
    col_of = lambda ci: _PALETTE[ci % len(_PALETTE)]
    drawn_faces = []                # (canonical DT, orbits) per row, for the verify pass

    for i, cl in enumerate(reps):
        canon = canonical_dt_string(cl["rep_dt"])

        # PUNCTURE ENUMERATION: the different plane pictures of a SINGLE diagram come
        # from turning different faces to the outside.  Rather than rely on which raw
        # codes happened to be sampled, enumerate the faces TIED for the largest
        # boundary -- exactly the faces draw_dt could pick as the outer 'puncture' --
        # render the canonical diagram with each, and keep the ones that draw
        # genuinely differently (rotation / flip / strand-reversal).  The canonical is
        # drawn with its default (canonical) puncture in the blue-outlined slot, so we
        # EXCLUDE that drawing; the remaining panels are the other distinct punctures.
        tie_sigs = _largest_tie_face_signatures(canon)
        orbits = _puncture_distinct_drawings(canon, K)
        n_shown = 1 + len(orbits)                       # canonical + distinct punctures
        capped = (len(orbits) >= K)                     # may be more distinct beyond the cap
        drawn_faces.append((canon, orbits))

        name = _diagram_name(canon)[0]
        head = ("Rank %d — %s" % (i + 1, name)) if name else ("Rank %d" % (i + 1))
        axt = fig.add_subplot(gs[i, 0])
        axt.axis("off")
        txt = ("%s\n(group #%d)\n%d largest-face tie%s\n%d total occurrences\n"
               "%s%d distinct drawing%s shown\n(incl. canonical)\n\ncanonical DT:\n%s"
               % (head, cl["rep_id"], len(tie_sigs), "" if len(tie_sigs) == 1 else "s",
                  cl["multiplicity"],
                  "≥" if capped else "", n_shown, "" if n_shown == 1 else "s",
                  "\n".join(textwrap.wrap(canon.replace("DT: ", ""), 24))))
        axt.text(0.0, 1.0, txt, transform=axt.transAxes, va="top", ha="left", fontsize=8.5)

        axc = fig.add_subplot(gs[i, 1])
        _draw_one(axc, canon, col_of, show_labels=False, rasterize=rasterize)
        # Mark the canonical panel with a Rectangle, NOT with the axes spines:
        # _render_draw ends with ax.axis("off"), and Matplotlib skips the spines
        # entirely on an axis-off axes, so re-showing them draws nothing at all.
        axc.add_patch(plt.Rectangle((0, 0), 1, 1, transform=axc.transAxes,
                                    fill=False, edgecolor="#2c7fb8", lw=2.0,
                                    clip_on=False, zorder=5))
        axc.set_title("canonical\n(default puncture)\n%s"
                      % "\n".join(textwrap.wrap(canon.replace("DT: ", ""), 24)),
                      fontsize=5.0 if i else 6.0)

        for j in range(K):
            ax = fig.add_subplot(gs[i, 2 + j])
            if j < len(orbits):
                rep = orbits[j]["rep"]
                # pinned by face IDENTITY, not by signature: two different faces can
                # share a crossing-ID signature, and the signature selector cannot
                # tell them apart
                _draw_one(ax, canon, col_of, show_labels=False, rasterize=rasterize,
                          puncture=rep["key"])
                cap = "puncture: " + "+".join(rep["sig"])
                extra = len(orbits[j]["members"]) - 1
                if extra:
                    cap += "   (+%d symmetric face%s)" % (extra, "" if extra == 1 else "s")
                ax.set_title("\n".join(textwrap.wrap(cap, 24)), fontsize=5.5)
            else:
                ax.axis("off")

    fig.suptitle("The plane drawings draw_dt can produce for each diagram, by which face it turns "
                 "to the outside\n"
                 "(each row = one canonical group; the panels are the genuinely different DRAWINGS "
                 "obtained by puncturing the faces TIED FOR THE LARGEST boundary -- the only faces "
                 "draw_dt ever chooses)",
                 fontsize=13, y=0.999)
    fig.text(0.5, 0.004,
             "Blue-outlined = canonical representative, drawn with its default puncture.  A diagram "
             "lives on a sphere and draw_dt must turn one face to the outside; when several faces tie "
             "for the largest boundary, each choice gives a different plane picture of the SAME diagram. "
             "Two punctures give the same panel exactly when a symmetry of the diagram carries one to "
             "the other -- rotation, in-plane mirror, component swap, strand reversal, or the mirror "
             "through the plane of the paper that flips every crossing at once (the same up-to-mirror "
             "convention the de-duplication uses).  This is NOT every plane drawing the diagram has: "
             "puncturing a smaller face gives more, which enumerate_puncturing_dt.py enumerates in full.",
             ha="center", va="bottom", fontsize=8, color="#555555")
    _strip_clip_paths(fig)                   # remove all clipping masks from the SVG
    fig.savefig(path, bbox_inches="tight", dpi=170 if rasterize else 100)
    plt.close(fig)
    if verify:
        verify_raw_grouping(drawn_faces)


def verify_raw_grouping(drawn_faces, log=print):
    """Check the panels of the raw-grouping figure against the geometry they claim.

    Two independent checks, in the two directions the grouping can fail:

    1. EQUIVARIANCE -- if a symmetry carries face A to face B they were merged, so
       their drawings must be congruent.  Every non-representative face is laid
       out and compared with its orbit's representative.  A mismatch means the
       layout is not equivariant under the diagram's symmetries (a layout bug).
    2. DUPLICATES -- no two panels in a row may be the same picture.  Checked from
       the rendered geometry alone, over every rigid alignment, including the
       over-strand direction at each crossing, so it can contradict the
       combinatorial grouping rather than restate it.

    This pair is what caught both bugs fixed in V2.2's predecessor: comparing
    nudged positions, and a grouping key blind to over/under.
    """
    bad_equiv, dupes, checked = [], [], 0
    for canon, orbits in drawn_faces:
        model = DDOL.build_model(DDOL.parse_dt(canon))
        G = DDOL.build_gadget_graph(model)
        pics = []
        for orb in orbits:
            P_exact, P_drawn = _positions_for_face(model, G, orb["rep"]["key"])
            want = _crossing_distance_key(model, P_exact)
            pics.append(_panel_picture(model, P_drawn))
            for m in orb["members"][1:]:
                checked += 1
                Q, _ = _positions_for_face(model, G, m["key"])
                if _crossing_distance_key(model, Q) != want:
                    bad_equiv.append("+".join(m["sig"]))
        n_cross = len(model["crossings"])
        for a in range(len(pics)):
            for b in range(a + 1, len(pics)):
                counts = _alignment_mismatches(pics[a], pics[b])
                if counts and (counts[0] == 0 or counts[-1] == n_cross):
                    dupes.append((canon, a + 1, b + 1))
    log("  [verify] raw figure: %d symmetric face%s laid out; %s"
        % (checked, "" if checked == 1 else "s",
           "every one congruent to its panel." if not bad_equiv
           else "EQUIVARIANCE MISMATCH: %s" % bad_equiv[:8]), flush=True)
    log("  [verify] raw figure: %s"
        % ("no two panels in a row are the same drawing."
           if not dupes else "DUPLICATE PANELS: %s" % dupes[:8]), flush=True)
    return {"checked": checked, "equivariance": bad_equiv, "duplicates": dupes}


# --------------------------------------------------------------------------- #
#  5. CLI
# --------------------------------------------------------------------------- #
DEFAULT_DT = "DT: [(-8,-12,16),(-24,-22,-28,-26),(-10,-14,-2),(-20,-6,-18,-4)]"


def launch_gui(defaults=None):
    """Tkinter front-end: fill in the parameters and press Run.  Launched when the
    script is started with no arguments or with --gui.  Falls back to a CLI run if
    Tkinter / a display is unavailable."""
    try:
        import tkinter as tk
        from tkinter import scrolledtext, filedialog, messagebox
        root = tk.Tk()                       # fails here if there is no display
    except Exception as exc:  # no Tk / no display
        print("Tkinter GUI unavailable (%s); running the pipeline on the CLI instead.\n" % exc)
        if defaults is not None:
            defaults.gui = False
            run_pipeline(defaults)
        return

    import threading
    import queue as _queue

    def dv(name, fallback):
        return str(getattr(defaults, name, fallback)) if defaults is not None else str(fallback)

    root.title("DT Diagram Scorer  —  score_diagramV2_5")
    frm = tk.Frame(root, padx=10, pady=8)
    frm.pack(fill="x")

    # GUI default for "Reset to root" is 20 (respects a non-zero value passed via --reset-every).
    reset_default = getattr(defaults, "reset_every", 0) or 20

    # output-file rows get a "Browse..." button (choose folder + file name via a save dialog)
    _save_dialog = {
        "checkpoint": [("JSONL checkpoint", "*.jsonl")],
        "xlsx": [("Excel workbook", "*.xlsx")],
        "svg": [("SVG figure", "*.svg"), ("PNG image", "*.png")],
        "raw_svg": [("SVG figure", "*.svg"), ("PNG image", "*.png")],
        "json": [("JSON", "*.json")],
    }

    def _make_browser(var, filetypes, defext, confirm=True):
        # confirm=False (checkpoint): pick an existing OR new file without the misleading
        # "…already exists, replace?" prompt, since a checkpoint is resumed, not overwritten.
        def _browse():
            path = filedialog.asksaveasfilename(
                title="Choose or open a checkpoint file" if not confirm else "Choose output location",
                defaultextension=defext,
                filetypes=filetypes + [("All files", "*.*")],
                initialfile=os.path.basename(var.get() or ""),
                initialdir=os.path.dirname(var.get() or "") or os.getcwd(),
                confirmoverwrite=confirm)
            if path:
                var.set(path)
        return _browse

    # per-field help shown by the light-blue "?" badge: (title, body-with-example)
    HELP = {
        "dt": ("DT code",
               "The signed Dowker–Thistlethwaite code of the STARTING diagram, grouped by "
               "component. A negative even number marks that the over-strand passes there. Every "
               "alternative the search finds is the SAME link, just drawn differently.\n\n"
               "Example:\nDT: [(-8,-12,16),(-24,-22,-28,-26),(-10,-14,-2),(-20,-6,-18,-4)]"),
        "rounds": ("Calls (simplifier runs)",
               "How many times to run the simplifier. In HARVESTING mode each call keeps every "
               "diagram it meets at the smallest crossing number — often a hundred or more — so "
               "far fewer calls are needed than the old one-per-round chain. In LEGACY CHAIN mode "
               "each call yields exactly one diagram.\n\n"
               "Example: 24 harvesting (6 was enough to get every diagram right on a 196-knot "
               "test); 999 for the legacy chain."),
        "backtrack_rounds": ("Backtrack rounds (per call)",
               "Inside each call, how many times the diagram is randomly re-tangled and "
               "re-simplified. This is where the diagrams actually come from: in harvesting mode "
               "every one of these cycles that ties the smallest crossing number is kept, so this "
               "number — not Calls — is the main yield lever.\n\nExample: 200."),
        "backtrack_steps": ("Backtrack steps (perturbation size)",
               "How many random crossing moves make up one re-tangle before re-simplifying. It "
               "matters more than it looks: on one test knot 30 steps found 12 distinct minimal "
               "diagrams where 20 steps found 7.\n\nExample: 30."),
        "reset_every": ("Reset every N rounds (legacy chain only)",
               "LEGACY CHAIN ONLY — greyed out in harvesting mode. Re-start the walk from the "
               "original DT every N rounds (0 = never), so the chain does not drift into one "
               "region.\n\n"
               "Harvesting does not need it: every call already restarts from a diagram at the "
               "smallest crossing count found so far, and switches away from the starting code "
               "automatically once something smaller turns up.\n\nExample: 20."),
        "seed": ("Seed (does NOT make a run repeatable)",
               "The search is randomized: each call randomly re-tangles the diagram before "
               "re-simplifying, which is how it finds different drawings. The seed steers the "
               "script's own choices, but it does NOT make a run repeatable — SnapPy/spherogram "
               "picks its moves from collections ordered by object identity, so the same seed "
               "gives different diagrams every time. Verified in fresh processes with "
               "PYTHONHASHSEED fixed.\n\n"
               "• To replay a run exactly, use the Checkpoint file — that is the only exact "
               "record.\n"
               "• Counts are therefore lower bounds. If an exact count matters, run several "
               "times and pool the results rather than making one run longer.\n\n"
               "Example: 20260708."),
        "verify": ("Verify sample",
               "An extra EXACT isomorphism check on the grouped diagrams, on top of the fast "
               "signature. 0 = off, -1 = check every member, N = check N per group. Use it to be "
               "certain no genuinely different diagrams were merged. Recomputed every run, even "
               "when resuming from a checkpoint.\n\nExample: 10."),
        "max_draw": ("Max diagrams drawn",
               "Cap on how many ranked representatives appear in the figure. Set it at or above the "
               "number of distinct diagrams the search finds to show them all.\n\nExample: 60."),
        "raw_max_per_class": ("Raw: distinct drawings per group",
               "In the raw-grouping figure, how many DISTINCT drawings to show per group "
               "(de-duplicated by rotation / flip / strand-reversal of the picture; the canonical is "
               "always one of them).\n\nExample: 20."),
        "write_raw": ("Write raw-grouping figure",
               "When ticked, also write the raw-grouping figure — it shows how the many raw sampled "
               "DRAWINGS collapse into the de-duplicated representatives (grouped in rank order, "
               "de-duplicated by rotation / flip / strand-reversal of the picture). Untick it to "
               "skip that figure; its path, per-group count and rasterize option then grey out."),
        "checkpoint": ("Checkpoint file",
               "A JSONL file (use Browse to choose its FOLDER and name) recording every diagram "
               "produced during generation — one line each. It stores only the generated codes, "
               "not scores or figures, and it is the ONLY exact record of a run, since the seed "
               "does not make one repeatable.\n\n"
               "• Resume / extend: if the file exists, generation continues from it. Asking for "
               "more calls adds only the difference.\n"
               "• Always recomputed: dedup, Verify, scoring and ALL figures are rebuilt fresh "
               "from the loaded codes every run — so changing Verify or the raw settings and "
               "re-running takes effect without re-generating.\n"
               "• The root DT must match the file, or the run is refused. Do not mix a harvesting "
               "checkpoint with a legacy-chain one: they hold different things.\n\n"
               "Example: /path/to/chain.jsonl."),
        "xlsx": ("Excel output",
               "Path for the metrics workbook (.xlsx); blank = skip. Contains every score, the "
               "diagram names + clasp structure, and a colour-coded metric-direction legend.\n\n"
               "Example: diagram_scores.xlsx."),
        "svg": ("Ranked-figure SVG",
               "Path for the main figure — each ranked diagram drawn four ways (labelled diagram, "
               "2-D skeleton, and 3-D sphere from two views); blank = skip. A .png path also works.\n\n"
               "Example: diagram_scores.svg."),
        "raw_svg": ("Raw grouping SVG",
               "Path for the figure showing how the many raw sampled DRAWINGS collapse into the "
               "de-duplicated representatives (de-duplicated by rotation / flip / strand-reversal of "
               "the picture only); blank = skip. A .png path also works.\n\nExample: grouping.svg."),
        "json": ("JSON output",
               "Path for the full machine-readable results (all metrics + any membership checks); "
               "blank = skip.\n\nExample: results.json."),
        "generator": ("Generator",
               "HARVESTING (default) keeps every diagram that ties the smallest crossing number "
               "inside each simplifier call. LEGACY CHAIN is the older behaviour: one diagram per "
               "call, the rest discarded.\n\n"
               "The difference is large. Checked against an exhaustive enumeration of every "
               "alternating knot of 10 crossings or fewer — 196 knots, 509 minimal diagrams known "
               "exactly — harvesting found 509 of 509, right on all 196. The legacy chain found "
               "438 in one run and 463 pooled over three, and running it 20x longer added "
               "nothing.\n\n"
               "Use the legacy chain only to reproduce an older result."),
        "check": ("Check DT codes (membership test)",
               "After the run, each DT code you paste here (one per line) is tested against the "
               "diagrams that were found: the log reports whether an equivalent diagram was sampled "
               "and which ranked representative it matches (and how often it occurred). Use it to "
               "ask 'is this particular diagram among the ones my search produced?'"),
    }

    def _help_badge(parent, key):
        title, body = HELP[key]
        lbl = tk.Label(parent, text=" ? ", fg="#08306b", bg="#add8e6",
                       font=("TkDefaultFont", 9, "bold"), cursor="hand2",
                       relief="raised", bd=1)
        lbl.bind("<Button-1>", lambda e, t=title, b=body: messagebox.showinfo(t, b))
        return lbl

    vars_ = {}
    widgets = {}          # key -> {"entry": Entry, "browse": Button|None}

    def _full_row(key, label, val, browse=False):
        v = tk.StringVar(value=str(val))
        vars_[key] = v
        row = tk.Frame(frm)
        row.pack(fill="x", pady=2)
        tk.Label(row, text=label, width=26, anchor="w").pack(side="left")
        ent = tk.Entry(row, textvariable=v)
        ent.pack(side="left", fill="x", expand=True, padx=4)
        _help_badge(row, key).pack(side="left", padx=(2, 4))
        btn = None
        if browse and key in _save_dialog:
            ft = _save_dialog[key]
            btn = tk.Button(row, text="Browse…",
                            command=_make_browser(v, ft, ft[0][1].lstrip("*"),
                                                  confirm=(key != "checkpoint")))
            btn.pack(side="left")
        widgets[key] = {"entry": ent, "browse": btn, "label": None}
        return v

    def _num_pair(spec_a, spec_b):
        # two single-number fields sharing one row (narrow entries)
        row = tk.Frame(frm)
        row.pack(fill="x", pady=2)
        for spec in (spec_a, spec_b):
            if spec is None:
                continue
            key, label, val = spec
            v = tk.StringVar(value=str(val))
            vars_[key] = v
            lab_w = tk.Label(row, text=label, width=16, anchor="w")
            lab_w.pack(side="left")
            ent = tk.Entry(row, textvariable=v, width=11)
            ent.pack(side="left", padx=(0, 2))
            _help_badge(row, key).pack(side="left", padx=(2, 18))
            widgets[key] = {"entry": ent, "browse": None, "label": lab_w}

    _full_row("dt", "DT code", dv("dt", DEFAULT_DT))

    # --- generator: harvesting (default) or the legacy one-per-call chain ---
    gen_var = tk.StringVar(value="chain" if getattr(defaults, "legacy_chain", False) else "harvest")
    gen_row = tk.Frame(frm)
    gen_row.pack(fill="x", pady=(6, 0))
    tk.Label(gen_row, text="Generator", width=26, anchor="w").pack(side="left")
    tk.Radiobutton(gen_row, text="Harvesting (recommended)", variable=gen_var,
                   value="harvest").pack(side="left")
    tk.Radiobutton(gen_row, text="Legacy chain", variable=gen_var,
                   value="chain").pack(side="left", padx=(8, 0))
    _help_badge(gen_row, "generator").pack(side="left", padx=6)

    _num_pair(("rounds", "Calls", dv("rounds", 24)),
              ("backtrack_rounds", "Backtrack rounds", dv("backtrack_rounds", 200)))
    _num_pair(("backtrack_steps", "Backtrack steps", dv("backtrack_steps", 30)),
              ("reset_every", "Reset every N", reset_default))
    _num_pair(("seed", "Seed", dv("seed", 20260708)),
              ("verify", "Verify 0/-1/N", dv("verify", 10)))
    _num_pair(("max_draw", "Max drawn", dv("max_draw", 60)),
              ("raw_max_per_class", "Raw per group", dv("raw_max_per_class", 20)))
    _full_row("checkpoint", "Checkpoint file", dv("checkpoint", "chainV2.jsonl"), browse=True)
    _full_row("xlsx", "Excel out (blank=skip)",
              getattr(defaults, "xlsx", "") or "diagram_scores.xlsx", browse=True)
    _full_row("svg", "Ranked SVG (blank=skip)",
              getattr(defaults, "svg", "") or "diagram_scores.svg", browse=True)

    # --- raw-grouping figure: a checkbox gates its path + options (dynamic greying) ---
    write_raw_var = tk.BooleanVar(value=bool((getattr(defaults, "raw_svg", "") or "").strip()))
    raw_chk_row = tk.Frame(frm)
    raw_chk_row.pack(fill="x", pady=(6, 0))
    tk.Checkbutton(raw_chk_row, text="Write raw-grouping figure",
                   variable=write_raw_var).pack(side="left")
    _help_badge(raw_chk_row, "write_raw").pack(side="left", padx=6)
    _full_row("raw_svg", "Raw grouping SVG",
              getattr(defaults, "raw_svg", "") or "", browse=True)
    _full_row("json", "JSON out (blank=skip)", getattr(defaults, "json", "") or "", browse=True)

    # --- Check DT codes: membership test (explanation now lives in its "?" badge) ---
    chk_hdr = tk.Frame(frm)
    chk_hdr.pack(fill="x", pady=(10, 0))
    tk.Label(chk_hdr, text="Check DT codes (optional, one DT per line)",
             anchor="w", font=("TkDefaultFont", 10, "bold")).pack(side="left")
    _help_badge(chk_hdr, "check").pack(side="left", padx=6)
    check_text = tk.Text(frm, width=64, height=4)
    check_text.pack(fill="x", pady=2)

    btns = tk.Frame(root, padx=10, pady=4)
    btns.pack(fill="x")
    run_btn = tk.Button(btns, text="Run")
    run_btn.pack(side="left")
    raw_raster_var = tk.BooleanVar(value=bool(getattr(defaults, "raw_raster", False)))
    raw_raster_chk = tk.Checkbutton(
        btns, text="rasterize raw-grouping diagrams (faster/smaller; default = vector)",
        variable=raw_raster_var)
    raw_raster_chk.pack(side="left", padx=14)
    tk.Button(btns, text="Quit", command=root.destroy).pack(side="right")

    # dynamic greying: the raw-grouping path + options are active only when the box is ticked
    def _sync_raw_state(*_):
        state = "normal" if write_raw_var.get() else "disabled"
        for k in ("raw_svg", "raw_max_per_class"):
            w = widgets.get(k, {})
            if w.get("entry") is not None:
                w["entry"].config(state=state)
            if w.get("browse") is not None:
                w["browse"].config(state=state)
        raw_raster_chk.config(state=state)
        if write_raw_var.get() and not vars_["raw_svg"].get().strip():
            vars_["raw_svg"].set("diagram_scores_grouping.svg")   # sensible default when enabled

    write_raw_var.trace_add("write", _sync_raw_state)
    _sync_raw_state()          # apply the initial (default = off) greying

    # dynamic greying: "Reset every N" only means anything for the legacy chain, and the
    # first field counts CALLS when harvesting but ROUNDS on the chain.
    def _sync_generator_state(*_):
        legacy = (gen_var.get() == "chain")
        w = widgets.get("reset_every", {})
        if w.get("entry") is not None:
            w["entry"].config(state="normal" if legacy else "disabled")
        if w.get("label") is not None:
            w["label"].config(fg="black" if legacy else "#999999")
        lw = widgets.get("rounds", {}).get("label")
        if lw is not None:
            lw.config(text="Rounds" if legacy else "Calls")
        if not legacy and vars_["rounds"].get().strip() in ("99", "999"):
            vars_["rounds"].set("24")        # chain-sized counts are wasteful when harvesting

    gen_var.trace_add("write", _sync_generator_state)
    _sync_generator_state()

    log = scrolledtext.ScrolledText(root, width=104, height=20, font=("Menlo", 9))
    log.pack(fill="both", expand=True, padx=10, pady=(4, 10))

    q = _queue.Queue()
    _done = object()          # sentinel: worker finished, re-enable the Run button

    class _QWriter:
        def write(self, s):
            q.put(s)

        def flush(self):
            pass

    def _poll():
        # All widget updates happen here on the main thread, driven by the queue.
        try:
            while True:
                item = q.get_nowait()
                if item is _done:
                    run_btn.config(state="normal")     # re-enable for the next run
                else:
                    log.insert("end", item)
                    log.see("end")
        except _queue.Empty:
            pass
        root.after(120, _poll)

    def _run():
        try:
            a = argparse.Namespace(
                dt=vars_["dt"].get().strip() or DEFAULT_DT,
                rounds=int(vars_["rounds"].get()),
                backtrack_rounds=int(vars_["backtrack_rounds"].get()),
                backtrack_steps=int(vars_["backtrack_steps"].get()),
                reset_every=int(vars_["reset_every"].get()),
                seed=int(vars_["seed"].get()),
                checkpoint=vars_["checkpoint"].get().strip() or "chainV2.jsonl",
                max_seconds=0.0, generate_only=False,
                max_draw=int(vars_["max_draw"].get()),
                verify=int(vars_["verify"].get() or 0),
                xlsx=vars_["xlsx"].get().strip() or None,
                svg=vars_["svg"].get().strip() or None,
                raw_svg=(vars_["raw_svg"].get().strip() or None) if write_raw_var.get() else None,
                raw_max_per_class=int(vars_["raw_max_per_class"].get() or 20),
                raw_raster=raw_raster_var.get(),
                raw_verify=True,
                json=vars_["json"].get().strip() or None,
                check=[ln.strip() for ln in check_text.get("1.0", "end").splitlines() if ln.strip()],
                check_file=None, gui=False,
                legacy_chain=(gen_var.get() == "chain"),
                reset_mode="origin", reset_reencode=False, reset_subprocess=False,
            )
        except ValueError as exc:
            q.put("Invalid parameter: %s\n" % exc)
            return
        run_btn.config(state="disabled")
        q.put("\n===== run started =====\n")
        # Import SnapPy now, in the MAIN thread: on first import cypari/cysignals installs
        # signal handlers, which is only allowed in the main thread.  The pipeline runs in a
        # worker thread below, where its own `import snappy` is then a cached no-op.
        try:
            import snappy  # noqa: F401
        except Exception as exc:  # noqa: BLE001
            q.put("(SnapPy not available: %s)\n" % exc)

        def _worker():
            old = sys.stdout
            sys.stdout = _QWriter()
            try:
                run_pipeline(a)
                print("\n===== done =====\n")
            except Exception:  # noqa: BLE001
                import traceback
                print("ERROR:\n" + traceback.format_exc())
            finally:
                sys.stdout = old
                q.put(_done)                           # signal main thread to re-enable

        threading.Thread(target=_worker, daemon=True).start()

    run_btn.config(command=_run)
    q.put("Set parameters and press Run. Long runs stream progress here.\n"
          "Tip: a checkpoint file lets you stop and resume; large 'Rounds' can take minutes.\n")
    root.after(120, _poll)
    root.mainloop()


def main(argv=None):
    raw = list(sys.argv[1:]) if argv is None else list(argv)
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dt", default=DEFAULT_DT)
    ap.add_argument("--rounds", type=int, default=24,
                    help="number of simplifier CALLS.  With harvesting far fewer "
                         "are needed than the old one-diagram-per-round chain")
    ap.add_argument("--backtrack-rounds", type=int, default=200)
    ap.add_argument("--backtrack-steps", type=int, default=30)
    ap.add_argument("--seed", type=int, default=20260708,
                    help="base seed for the per-round seeds.  NOTE: this does not make "
                         "the chain repeatable (see the module docstring); use "
                         "--checkpoint to replay a chain, and pool several runs when an "
                         "exact count matters")
    ap.add_argument("--checkpoint", default="chainV2.jsonl")
    ap.add_argument("--canonical-limit", type=float, default=CANONICAL_COST_LIMIT,
                    help="skip the exact canonical form when it would need more than this "
                         "many relabellings (0 = never skip, and risk hours per class on "
                         "many-component links).  Default %d" % CANONICAL_COST_LIMIT)
    ap.add_argument("--legacy-chain", action="store_true",
                    help="use the V2.3 generator, which keeps ONE diagram per simplifier "
                         "call.  The default harvests every minimum-crossing diagram each "
                         "call and finds far more (see the module docstring)")
    ap.add_argument("--reset-mode", choices=list(RESET_MODES), default="origin",
                    help="where a reset restarts from: 'origin' the starting diagram "
                         "(old behaviour), 'equal' a uniformly random diagram among those "
                         "found so far, 'inverse' one weighted by 1/(1+times used). "
                         "Only has an effect together with --reset-every")
    ap.add_argument("--reset-reencode", action="store_true",
                    help="on each reset, hand spherogram a DIFFERENT ENCODING of the "
                         "chosen restart diagram (same diagram, new base points and "
                         "component order).  Changes spherogram's internal ordering, so "
                         "the restart explores differently.  Costs microseconds.  "
                         "LEGACY-CHAIN ONLY")
    ap.add_argument("--reset-subprocess", action="store_true",
                    help="on each reset, run that round in a FRESH interpreter.  The "
                         "strongest decorrelation available, but costs a process spawn "
                         "(~1-3 s) per reset, so use it with a large --reset-every.  "
                         "LEGACY-CHAIN ONLY")
    ap.add_argument("--reset-every", type=int, default=0,
                    help="re-root the chain at the original DT after every N rounds "
                         "(0 = never); avoids getting trapped cycling among a few diagrams")
    ap.add_argument("--max-seconds", type=float, default=0.0,
                    help="stop generation after this many seconds (0 = no limit); resumable")
    ap.add_argument("--generate-only", action="store_true")
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--svg", default=None)
    ap.add_argument("--raw-svg", default=None,
                    help="also write an SVG showing how the raw sampled diagrams group into "
                         "the de-duplicated representatives (rotation/flip equivalence)")
    ap.add_argument("--raw-max-per-class", type=int, default=20,
                    help="max distinct drawings shown per group in --raw-svg")
    ap.add_argument("--no-raw-verify", dest="raw_verify", action="store_false",
                    help="skip the raw-grouping figure's self-check (equivariance of the "
                         "merged punctures, and a geometric scan for duplicate panels)")
    ap.add_argument("--raw-raster", action="store_true",
                    help="rasterize the diagram panels in --raw-svg (faster/smaller SVG; "
                         "default is full vector art, which is slower to write for many diagrams)")
    ap.add_argument("--json", default=None)
    ap.add_argument("--max-draw", type=int, default=60)
    ap.add_argument("--check", action="append", default=[],
                    help="a DT code to test for membership among the sampled diagrams "
                         "(repeatable)")
    ap.add_argument("--check-file", default=None,
                    help="file with one DT code per line to test for membership")
    ap.add_argument("--verify", type=int, default=10, metavar="N",
                    help="confidence check: exact VF2 of each class rep against up to N "
                         "other distinct members (0 = off; use a large N or -1 for all)")
    ap.add_argument("--gui", action="store_true",
                    help="launch the graphical interface (also the default when no "
                         "arguments are given)")
    args = ap.parse_args(raw)
    globals()["CANONICAL_COST_LIMIT"] = int(getattr(args, "canonical_limit", CANONICAL_COST_LIMIT) or 0)

    if (not raw) or args.gui:
        launch_gui(args)
        return
    run_pipeline(args)


def run_pipeline(args, log=None):
    """Run generation -> dedup -> (verify) -> membership check -> score -> outputs.
    Prints progress with print(); the GUI redirects stdout to capture it."""
    t_start = time.time()
    harvest = not getattr(args, "legacy_chain", False)
    print("Generating (%s): %d calls, backtrack %dx%d ..."
          % ("harvesting" if harvest else "legacy chain", args.rounds,
             args.backtrack_rounds, args.backtrack_steps), flush=True)
    if harvest:
        chain = generate_archive(
            args.dt, args.rounds, args.backtrack_rounds, args.backtrack_steps,
            args.seed, checkpoint=args.checkpoint,
            max_seconds=(args.max_seconds or None),
        )
    else:
        chain = generate_chain(
            args.dt, args.rounds, args.backtrack_rounds, args.backtrack_steps,
            args.seed, checkpoint=args.checkpoint,
            max_seconds=(args.max_seconds or None), reset_every=args.reset_every,
            reset_mode=getattr(args, "reset_mode", "origin"),
            reset_reencode=getattr(args, "reset_reencode", False),
            reset_subprocess=getattr(args, "reset_subprocess", False),
        )
    done = len(chain) - 1
    if harvest:
        print("Harvested %d minimal DT codes." % len(chain), flush=True)
        if args.generate_only:
            return
    else:
        print("Chain has %d/%d rounds (%d DT codes)." % (done, args.rounds, len(chain)), flush=True)
        if args.generate_only or done < args.rounds:
            if done < args.rounds:
                print("Not finished; re-run to resume from checkpoint %s." % args.checkpoint)
            return

    print("Deduplicating %d DT codes ..." % len(chain), flush=True)
    classes = dedup(chain)
    print("  -> %d distinct representative diagrams." % len(classes), flush=True)

    if args.verify:
        n_sample = 0 if args.verify < 0 else args.verify
        rep = verify_classes(classes, sample=n_sample)
        bad = [r for r in rep if not r["all_isomorphic"]]
        print("Verification (exact VF2, sample=%s): %s"
              % (n_sample or "ALL", "all classes internally consistent"
                 if not bad else "MERGE ERROR in classes %s" % [r["rep_id"] for r in bad]),
              flush=True)
        for r in rep:
            print("  class #%d: %d checked, all isomorphic=%s"
                  % (r["rep_id"], r["checked"], r["all_isomorphic"]), flush=True)

    queries = list(args.check)
    if args.check_file and os.path.exists(args.check_file):
        with open(args.check_file) as fh:
            queries += [ln.strip() for ln in fh if ln.strip()]
    check_results = check_sampled(classes, queries) if queries else []
    if check_results:
        print("Membership check (%d queries):" % len(check_results), flush=True)
        for r in check_results:
            if r.get("error"):
                print("  [error] %s : %s" % (r["dt"], r["error"]))
            elif r["sampled"]:
                print("  SAMPLED   -> rep #%d (mult %d, first seen round %d) : %s"
                      % (r["matches_rep_id"], r["multiplicity"], r["first_round"], r["dt"]))
            else:
                print("  NOT found -> %s" % r["dt"])

    print("Scoring representatives ...", flush=True)
    scored = score_representatives(classes)

    # Pre-compute the true 3-D point group once per representative (shared by the
    # xlsx "3D point group" column and the SVG symmetry overlays), in each rep's
    # own 3-D frame so the drawn axes/planes/inversion dot are correctly oriented.
    for m in scored:
        try:
            _compute_sym3d(m["dt"], m["sphere3d"]["_centers3d"])
        except Exception:  # noqa: BLE001
            pass

    fps = set(m["linking_fp"] for m in scored if m["linking_fp"] is not None)
    if not fps:
        print("  same-link check: SKIPPED (no linking numbers available -- is SnapPy "
              "importable?)", flush=True)
    else:
        print("  same-link check: %d distinct linking-number fingerprint(s) among "
              "representatives%s (expected 1: all are the same link by construction)"
              % (len(fps), "" if len(fps) == 1 else "  <-- MISMATCH"), flush=True)

    # Raw-grouping figure: built AFTER scoring so its groups appear in the SAME order
    # (by composite rank) as the ranked figure, and are labelled with rank + name.
    if getattr(args, "raw_svg", None):
        make_raw_grouping_figure([m["_class"] for m in scored], args.raw_svg,
                                 max_per_class=getattr(args, "raw_max_per_class", 6),
                                 rasterize=getattr(args, "raw_raster", False),
                                 verify=getattr(args, "raw_verify", True))
        print("wrote %s" % args.raw_svg, flush=True)

    run_info = {
        "software": "score_diagramV2_5.py",
        "root_DT": args.dt,
        "rounds": args.rounds,
        "backtrack_rounds": args.backtrack_rounds,
        "backtrack_steps": args.backtrack_steps,
        "reset_every": args.reset_every,
        "canonical_cost_limit": CANONICAL_COST_LIMIT,
        "canonical_fallbacks": len(_COMPLEXITY_WARNED),
        "generator": "harvesting (every minimum-crossing diagram kept)"
                     if not getattr(args, "legacy_chain", False) else "legacy chain",
        "reset_mode": getattr(args, "reset_mode", "origin"),
        "reset_reencode": getattr(args, "reset_reencode", False),
        "reset_subprocess": getattr(args, "reset_subprocess", False),
        "seed": args.seed,
        "total_DT_codes": len(chain),
        "distinct_representatives": len(classes),
        "distinct_linking_fingerprints": len(fps),
        "jones_polynomial": "populated only under Sage (SnapPy standalone cannot compute it)",
        "scored_on": "canonical DT of each class (labeling-independent, reproducible)",
        "dedup_equivalence": "signed diagram isomorphism (rotation/reflection/relabel/"
                             "cyclic-permutation/component-reorder; over-under preserved). "
                             "Fast signature = WL hash + strand-length + face-degree spectra; "
                             "exact backstops: --verify (VF2) and canonical_dt().",
        "scoring_layout_2d": "Tutte (unit-circle boundary)",
        "figure_layout_2d": "shaped-tutte, ellipse, aspect 1.0 (draw_dt panel and skeleton share it)",
        "layout_3d": "spherical Kamada-Kawai crossing centers (unit sphere)",
        "generated_utc": time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()),
        "runtime_seconds": round(time.time() - t_start, 1),
    }

    if args.json:
        payload = {
            "run_info": run_info,
            "membership_check": check_results,
            "representatives": [
                {
                    "rep_id": m["_class"]["rep_id"],
                    "rank": m["rank"],
                    "dt": m["dt"],
                    "multiplicity": m["_class"]["multiplicity"],
                    "first_round": m["_class"]["rep_round"],
                    "jones": m["jones"],
                    "linking_fingerprint": list(m["linking_fp"]) if m["linking_fp"] else None,
                    "composite": m["composite"],
                    "quality": m["quality"],
                    "combinatorial": _strip_private({"c": m["combinatorial"]})["c"],
                    "graph": {k: v for k, v in m["graph"].items() if not k.startswith("_")},
                    "geom2d": {k: v for k, v in m["geom2d"].items() if not k.startswith("_")},
                    "sphere3d": {k: v for k, v in m["sphere3d"].items() if not k.startswith("_")},
                }
                for m in scored
            ],
        }
        with open(args.json, "w") as fh:
            json.dump(payload, fh, indent=2)
        print("wrote %s" % args.json)

    if args.xlsx:
        write_excel(scored, args.xlsx, run_info)
        print("wrote %s" % args.xlsx)
    if args.svg:
        make_figure(scored, args.svg, max_draw=args.max_draw)
        print("wrote %s" % args.svg)

    print("\nTop 3 representatives by composite:")
    for m in scored[:3]:
        print("  #%d  composite %.3f  strands %s  |Aut| %d  mult %d"
              % (m["_class"]["rep_id"], m["composite"],
                 m["combinatorial"]["strand_visit_lengths"],
                 m["graph"]["automorphism_order"], m["_class"]["multiplicity"]))


if __name__ == "__main__":
    main()
